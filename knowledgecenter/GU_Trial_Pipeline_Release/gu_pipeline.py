#!/usr/bin/env python3
"""
GU Oncology SoCal — Extended Trial Data Pull
=============================================
Version : 2.1.0
Author  : Sophie Zaaijer PhD (pipeline) / Claude (implementation)

New in v2.1:
  - Output restructured to one dated folder per run containing 5 Excel files
    (one per cancer type group + updates/analytics file)
  - Adrenal tumours added (adrenocortical carcinoma, pheochromocytoma)
  - All four NCCN taxonomies active in Trial Finder (Prostate, Bladder, Kidney, Testicular)
  - studyFirstPostDate captured for time-window analytics
  - Delta/updates file: changes vs previous run, stats by hospital, new-trial windows

Output (output/YYYY-MM-DD_HHMM/ folder, timestamped):
  prostate_trials_YYYY-MM-DD.xlsx         ← Trial Finder · Sites×PI · Eligibility · Outcomes
  bladder_trials_YYYY-MM-DD.xlsx
  kidney_adrenal_trials_YYYY-MM-DD.xlsx
  testicular_trials_YYYY-MM-DD.xlsx
  updates_YYYY-MM-DD.xlsx                 ← Changes · Stats · New-trial windows · Metadata
  all_trials_YYYY-MM-DD.csv              ← Flat file used for delta in next run

Usage:
  python3 gu_pipeline.py

Requirements:
  pip install requests pandas openpyxl
"""

from __future__ import annotations

import re
import sys
import time
import threading
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from pathlib import Path

import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Disease classifier ───────────────────────────────────────────────────────
sys.path.insert(0, str(Path(__file__).parent))
from nccn_classifier import (
    classify_trial,
    get_disease_settings_in_order,
    get_section_colors,
    get_nccn_stamp,
)

# ── Version + metadata ────────────────────────────────────────────────────────
VERSION   = "2.1.0"
PULLED_BY = "Sophie Zaaijer PhD"

# ── Paths ─────────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent
OUT  = ROOT / "output"
OUT.mkdir(exist_ok=True)

# ── API config (identical to master pipeline) ─────────────────────────────────
BASE_URL   = "https://clinicaltrials.gov/api/v2/studies"
GEO_FILTER = "distance(33.8,-118.0,180mi)"
STATUSES   = ["RECRUITING"]
STUDY_TYPE = "INTERVENTIONAL"
PAGE_SIZE  = 100
MAX_WORKERS = 5

CONDITIONS = [
    "prostate cancer",
    "prostate adenocarcinoma",   # catches trials where sponsor uses specific histology term
    "bladder cancer",
    "urothelial carcinoma",
    "renal cell carcinoma",
    "kidney cancer",
    "testicular cancer",
    "penile cancer",
    "upper tract urothelial",
    "adrenocortical carcinoma",
    "pheochromocytoma",
]

TARGET_INSTITUTIONS: set[str] = {
    "UCLA",
    "USC / Norris",
    "USC / LAC+USC Medical Center",
    "UCSD / Moores",
    "UC Irvine / Chao",
    "City of Hope",
    "City of Hope – Orange County",
    "City of Hope – Corona",
    "Hoag",
    "Cedars-Sinai",
    "Providence / St. Jude",
    "Providence Medical Foundation",
    "Loma Linda",
    "Loma Linda VA",
    "VA Greater Los Angeles",
    "VA Long Beach",
    "Scripps",
    "Sharp",
}


# ── Facility normalisation ────────────────────────────────────────────────────
_NORM_RULES: list[tuple[str, str]] = [
    (r'city of hope.*(irvine|lennar|orange county|huntington beach)|lennar.*city of hope',
     "City of Hope – Orange County"),
    (r'city of hope.*corona',                                        "City of Hope – Corona"),
    (r'city of hope|beckman research.*city of hope|duarte cancer',   "City of Hope"),
    (r'ucla|jonsson comprehensive|university of california.{0,6}los an[g]?[e]?les'
     r'|david geffen school of medicine|westwood cancer|administrative address.*ucla',
     "UCLA"),
    (r'los angeles county.{0,5}usc|los angeles general medical|lac\+usc',
     "USC / LAC+USC Medical Center"),
    (r'usc|norris|univeristy of southern california|university of southern california'
     r'|keck medicine of usc|koman family outpatient|institute of urology.*southern california'
     r'|university of south california',
     "USC / Norris"),
    (r'ucsd|uc san diego|university of california.{0,6}san diego'
     r'|moores cancer center|rebecca and john moores',               "UCSD / Moores"),
    (r'uc irvine|uci |uci health|university of california.{0,6}irvine'
     r'|chao family comprehensive|irvine medical center',            "UC Irvine / Chao"),
    (r'cedars?.sin[ae]i|cedars?.senai|angeles clinic',               "Cedars-Sinai"),
    (r'veterans affairs loma linda|va.*loma linda',                  "Loma Linda VA"),
    (r'loma linda',                                                   "Loma Linda"),
    (r'greater los angeles.*(va|veterans)|(va|veterans).{0,10}greater los angeles',
     "VA Greater Los Angeles"),
    (r'va long beach|long beach va|tibor rubin',                     "VA Long Beach"),
    (r'hoag',                                                         "Hoag"),
    (r'scripps',                                                      "Scripps"),
    (r'kaiser permanente|kaiser.permanente',                          "Kaiser Permanente"),
    (r'providence.*st\.? jude|st\.? joseph heritage|virginia k\.? crosson'
     r'|john wayne cancer|saint john.*cancer',                       "Providence / St. Jude"),
    (r'providence',                                                   "Providence Medical Foundation"),
    (r'sharp',                                                        "Sharp"),
]
_NORM_COMPILED = [(re.compile(p, re.IGNORECASE), lbl) for p, lbl in _NORM_RULES]


def normalize_facility(facility: str) -> str:
    for rx, lbl in _NORM_COMPILED:
        if rx.search(facility or ""):
            return lbl
    return facility or "Unknown"


_SOCAL_CITIES = {
    "los angeles","santa monica","beverly hills","west hollywood","el segundo",
    "torrance","long beach","pasadena","duarte","pomona","west covina","covina",
    "glendale","burbank","arcadia","downey","whittier",
    "san diego","la jolla","chula vista","el cajon","encinitas","escondido",
    "national city","oceanside","santee","san marcos",
    "irvine","orange","newport beach","anaheim","fullerton","santa ana",
    "garden grove","mission viejo","laguna hills","laguna niguel","aliso viejo",
    "tustin","costa mesa","huntington beach","fountain valley",
    "loma linda","riverside","san bernardino","ontario","rancho cucamonga",
    "fontana","redlands","murrieta","temecula","corona","upland",
    "ventura","oxnard","thousand oaks","camarillo",
}
_SOCAL_KW = [
    "UCLA","USC","Keck","UCSD","Moores","UC Irvine","City of Hope",
    "Cedars-Sinai","Cedars Sinai","Loma Linda","Hoag","Scripps","Sharp",
    "Providence","Kaiser Permanente",
]


def is_socal_site(city: str, facility: str) -> bool:
    if (city or "").lower() in _SOCAL_CITIES:
        return True
    fl = (facility or "").lower()
    return any(kw.lower() in fl for kw in _SOCAL_KW)


# ── PI name cleaning ──────────────────────────────────────────────────────────
_NON_HUMAN_RE = re.compile(
    r'director|pfizer|janssen|ctlilly|abbvie|squibb|novartis|roche|merck|'
    r'astrazeneca|bristol.myers|bms\b|sanofi|bayer|'
    r'global\s+lead|global\s+compliance|global\s+leader|'
    r'clinical\s+lead|clinical\s+leader|clinical\s+sites|clinical\s+trials?[,\s]|'
    r'clinical\s+management|medical\s+lead|medical\s+monitor|medical\s+officer|'
    r'chief\s+officer|head\s+of\s+science|'
    r'\bofficer\b|\bmonitor\b|\bcompliance\b|'
    r'^\s*lead\s*[,\s]|^\s*management\s*[,\s]|^\s*sites\s*[,\s]|'
    r'revolution medicines|janux therapeutics|iambic therapeutics|'
    r'indaptus therapeutics|arsenal biosciences|evolveimmune|adaptimmune|'
    r'full.life gmbh|riboscience|idera\b|gilead\b|iovance|'
    r'johnson\s+&?\s*johnson|\btrial[,\s]|\bteam[,\s]|'
    r'pharmaceuticals?[,\s]|therapeutics[,\s]|biosciences?[,\s]|medicines?[,\s]|'
    r'\binc[,.\s]|\bllc\b|\bgmbh\b|'
    r'study\s+director|site\s+director|'
    r'call\s+1-|1-877|1-317|'
    r'pharm\.?d\b|pharmd\b|'
    r'\boncology\s*$|\bhematology\s*$|\boncologist\b',
    re.IGNORECASE,
)
_CRED_RE = re.compile(
    r'^(m\.?d\.?|ph\.?d\.?|d\.?o\.?|m\.?p\.?h\.?|m\.?s\.?|r\.?n\.?|'
    r'n\.?p\.?|p\.?a\.?|f\.?a\.?c\.?[spo]\.?|f\.?r\.?c\.?p\.?c?\.?|'
    r'b\.?c\.?|m\.?b\.?a\.?|pharm\.?d\.?|prof\.?|bsn|cnmt|psyd|rph|pharmd)$',
    re.IGNORECASE,
)
_SUFFIX_RE = re.compile(r'^(jr\.?|sr\.?|ii|iii|iv)$', re.IGNORECASE)
_COMMON_GIVEN = {
    'james','john','robert','michael','william','david','richard','joseph',
    'thomas','charles','mark','paul','steven','andrew','kenneth','matthew',
    'brian','eric','jason','ryan','scott','adam','christopher','peter',
    'jonathan','nicholas','joshua','benjamin','daniel','kevin','gary','alan',
    'tyler','bobby','jeffrey','derek','brett','gregory','patrick','phillip',
    'raymond','stephen','timothy','walter','wayne','zachary',
    'emily','sarah','jessica','ashley','mary','jennifer','linda','barbara',
    'patricia','elizabeth','susan','karen','nancy','lisa','margaret','betty',
    'tanya','andrea','heather','virginia','monica','rose','viola','jacqueline',
    'anna','erin','laura','amanda','stephanie','rebecca','angela','nicole',
    'melissa','rachel','danielle','lauren','amy','cynthia','diana','donna',
    'rana','neeraj','amar','jeremie','hideki','sarmad','felix','sandip',
    'atish','koho','alain','sumanta','siamak','bertram','maha','arash',
    'karim','milan','hamid','mehrdad','ali','xiao','shilpa','nazli',
}


def _is_non_human(name: str) -> bool:
    return not name or bool(_NON_HUMAN_RE.search(name))


def _clean_pi_name(raw: str) -> str:
    if not raw or not raw.strip():
        return "Not listed"
    n = raw.strip().rstrip(",;./")
    parts   = [p.strip() for p in n.split(",")]
    cleaned = [p for p in parts if p and not _CRED_RE.match(p)]
    n = ", ".join(cleaned).strip().rstrip(",;. ")
    n = re.sub(r"\s{2,}", " ", n)
    if not n or _is_non_human(n):
        return "Not listed"
    cp = [p.strip() for p in n.split(",") if p.strip()]
    if len(cp) >= 2:
        lw = cp[0].split()[0].lower() if cp[0] else ""
        rw = cp[1].split()[0].lower() if cp[1] else ""
        if lw in _COMMON_GIVEN and rw not in _COMMON_GIVEN:
            n = f"{cp[1]}, {cp[0]}"
    else:
        words = n.split()
        if len(words) >= 2:
            if len(words) >= 3 and _SUFFIX_RE.match(words[-1]):
                n = f"{words[-2]} {words[-1]}, {' '.join(words[:-2])}"
            else:
                n = f"{words[-1]}, {' '.join(words[:-1])}"
    return n or "Not listed"


def canonical_pi(name: str) -> str:
    if not name or name == "Not listed":
        return ""
    if "," in name:
        last, rest = name.split(",", 1)
        fi = rest.strip()[0].lower() if rest.strip() else ""
        return f"{last.strip().lower()}|{fi}"
    words = name.split()
    last  = words[-1].lower() if words else ""
    fi    = words[0][0].lower() if len(words) > 1 else (last[0] if last else "")
    return f"{last}|{fi}"


# ── Cancer type classification ────────────────────────────────────────────────
_CANCER_RE = [
    ("Prostate",           re.compile(r'prostat', re.IGNORECASE)),
    ("Bladder/Urothelial", re.compile(r'bladder|urothelial|upper.tract|transitional.cell', re.IGNORECASE)),
    ("Kidney/RCC",         re.compile(r'renal|kidney|nephr|renal.cell|RCC', re.IGNORECASE)),
    ("Adrenal",            re.compile(r'adrenal|adrenocortical|pheochromocytoma|paraganglioma', re.IGNORECASE)),
    ("Testicular/GCT",     re.compile(r'testicular|testis|germ.cell|seminoma|\\bNSGCT\\b|\\bGCT\\b', re.IGNORECASE)),
    ("Penile",             re.compile(r'penile|penis', re.IGNORECASE)),
]


def classify_cancer_types(conditions: str, title: str = "") -> list[str]:
    full = (conditions or "").lower()
    matched = [ct for ct, rx in _CANCER_RE if rx.search(full)]
    if not matched:
        t = (title or "").lower()
        matched = [ct for ct, rx in _CANCER_RE if rx.search(t)]
    return matched if matched else ["Basket"]


# ── PI extraction ─────────────────────────────────────────────────────────────
def extract_pi(study: dict) -> tuple[str, str]:
    ps        = study.get("protocolSection", {})
    officials = ps.get("contactsLocationsModule", {}).get("overallOfficials", [])
    for o in officials:
        if "principal" in o.get("role", "").lower():
            name = _clean_pi_name(o.get("name", ""))
            if name != "Not listed":
                return name, o.get("affiliation", "")
    rp = ps.get("sponsorCollaboratorsModule", {}).get("responsibleParty", {})
    if rp.get("type") == "PRINCIPAL_INVESTIGATOR":
        name = _clean_pi_name(rp.get("investigatorFullName", ""))
        if name != "Not listed":
            return name, rp.get("investigatorAffiliation", "")
    return "Not listed", ""


def extract_site_contact(loc: dict) -> tuple[str, str, str]:
    """
    Return (pi_name, pi_email, pi_phone) for a single location.

    CT.gov v2 API stores contacts in two places per location:
      1. loc["contacts"] — list of {name, role, phone, email}
         role = "PRINCIPAL_INVESTIGATOR" | "CONTACT" | "BACKUP"
      2. loc["investigators"] (present in some studies) — list of {name, role}

    Priority: PRINCIPAL_INVESTIGATOR contact with email > PRINCIPAL_INVESTIGATOR
    without email > any named CONTACT with email > global fallback (caller's job).
    """
    pi_name = pi_email = pi_phone = ""

    # Pass 1: look for an explicit PRINCIPAL_INVESTIGATOR contact
    for c in loc.get("contacts", []):
        if c.get("role", "").upper() == "PRINCIPAL_INVESTIGATOR":
            name = _clean_pi_name(c.get("name", ""))
            if name and name != "Not listed":
                pi_name  = name
                pi_email = c.get("email", "")
                pi_phone = c.get("phone", "")
                break   # take the first PI found

    # Pass 2: fall back to investigators list (present in some CT.gov v2 responses)
    if not pi_name:
        for inv in loc.get("investigators", []):
            if "principal" in inv.get("role", "").lower():
                name = _clean_pi_name(inv.get("name", ""))
                if name and name != "Not listed":
                    pi_name = name
                    break

    # Pass 3: if still no PI, take first named CONTACT (site coordinator) for the email/phone
    if not pi_email and not pi_phone:
        for c in loc.get("contacts", []):
            if c.get("name") and (c.get("email") or c.get("phone")):
                pi_email = pi_email or c.get("email", "")
                pi_phone = pi_phone or c.get("phone", "")
                break

    return pi_name, pi_email, pi_phone


# Keep legacy alias so older call sites compile without error
def extract_site_pi(loc: dict) -> str:
    name, _, _ = extract_site_contact(loc)
    return name


# ── Eligibility parser ────────────────────────────────────────────────────────
def parse_eligibility(text: str) -> tuple[str, str]:
    """
    Split CT.gov eligibility blob into (inclusion, exclusion).
    CT.gov format:
        Inclusion Criteria:
          * item
        Exclusion Criteria:
          * item
    """
    if not text:
        return "", ""
    excl = re.search(r'\bexclusion criteria\s*:?\s*\n', text, re.IGNORECASE)
    if excl:
        incl_raw = text[:excl.start()].strip()
        excl_raw = text[excl.end():].strip()
    else:
        incl_raw = text.strip()
        excl_raw = ""
    # Strip "Inclusion Criteria:" header
    incl_raw = re.sub(
        r'^\s*inclusion criteria\s*:?\s*\n?', '', incl_raw, flags=re.IGNORECASE
    ).strip()
    return incl_raw, excl_raw


# ── Extended study parser ─────────────────────────────────────────────────────
def parse_study_extended(study: dict) -> tuple[dict | None, list[dict]]:
    """
    Returns (study_record, site_rows).

    study_record : one dict — all study-level fields (trial details + eligibility + outcomes).
                   None if no qualifying SoCal sites found.
    site_rows    : list — one dict per qualifying SoCal site at a target institution.
    """
    ps = study.get("protocolSection", {})

    id_mod       = ps.get("identificationModule",        {})
    status_mod   = ps.get("statusModule",                {})
    design_mod   = ps.get("designModule",                {})
    sponsor_mod  = ps.get("sponsorCollaboratorsModule",  {})
    desc_mod     = ps.get("descriptionModule",           {})
    conds_mod    = ps.get("conditionsModule",            {})
    arms_mod     = ps.get("armsInterventionsModule",     {})
    outcomes_mod = ps.get("outcomesModule",              {})
    elig_mod     = ps.get("eligibilityModule",           {})
    contacts_mod = ps.get("contactsLocationsModule",     {})

    # ── Identification ────────────────────────────────────────────────────
    nct_id = id_mod.get("nctId", "")
    title  = id_mod.get("officialTitle") or id_mod.get("briefTitle", "")

    sec_ids = "; ".join(
        s.get("id", "")
        for s in id_mod.get("secondaryIdInfos", [])
        if s.get("id")
    )

    # ── Status / dates ────────────────────────────────────────────────────
    status          = status_mod.get("overallStatus", "")
    start_date      = status_mod.get("startDateStruct",              {}).get("date", "")
    end_date        = status_mod.get("primaryCompletionDateStruct",  {}).get("date", "")
    last_update     = status_mod.get("lastUpdatePostDateStruct",     {}).get("date", "")
    first_post_date = status_mod.get("studyFirstPostDateStruct",     {}).get("date", "")

    # ── Design ────────────────────────────────────────────────────────────
    phase      = ", ".join(design_mod.get("phases", []) or ["N/A"])
    study_type = design_mod.get("studyType", "")

    di    = design_mod.get("designInfo", {})
    mi    = di.get("maskingInfo", {})
    masking_detail = di.get("masking", "") or mi.get("masking", "")
    who   = ", ".join(mi.get("whoMasked", []))
    design_str = " | ".join(filter(None, [
        f"Allocation: {di.get('allocation','')}"          if di.get("allocation") else "",
        f"Model: {di.get('interventionModel','')}"        if di.get("interventionModel") else "",
        f"Purpose: {di.get('primaryPurpose','')}"         if di.get("primaryPurpose") else "",
        (f"Masking: {masking_detail}" + (f" ({who})" if who else "")) if masking_detail else "",
    ]))

    # ── Enrollment ────────────────────────────────────────────────────────
    enroll       = design_mod.get("enrollmentInfo", {})
    enroll_count = enroll.get("count", "")
    enroll_type  = enroll.get("type", "")
    enroll_est   = str(enroll_count) if enroll_type == "ESTIMATED" else ""
    enroll_act   = str(enroll_count) if enroll_type == "ACTUAL"    else ""

    # ── Sponsors ──────────────────────────────────────────────────────────
    lead_sponsor        = sponsor_mod.get("leadSponsor", {}).get("name", "")
    collabs             = sponsor_mod.get("collaborators", [])
    additional_sponsors = "; ".join(c.get("name", "") for c in collabs if c.get("name"))

    rp       = sponsor_mod.get("responsibleParty", {})
    rp_type  = rp.get("type", "")
    rp_name  = rp.get("investigatorFullName", "")
    rp_affil = rp.get("investigatorAffiliation", "")
    if rp_type == "SPONSOR":
        info_provided_by = f"Sponsor ({lead_sponsor})"
    elif rp_name:
        info_provided_by = rp_name + (f", {rp_affil}" if rp_affil else "")
    else:
        info_provided_by = rp_type or lead_sponsor

    # ── Description ───────────────────────────────────────────────────────
    brief_summary = (desc_mod.get("briefSummary") or "").strip()

    # ── Conditions ────────────────────────────────────────────────────────
    conditions_list = conds_mod.get("conditions", [])
    conditions_full = "; ".join(conditions_list)

    # ── Interventions ─────────────────────────────────────────────────────
    intervention_str = "; ".join(
        f"{iv.get('type','Other')}: {iv.get('name','')}"
        for iv in arms_mod.get("interventions", [])
        if iv.get("name")
    )

    # ── Outcomes ──────────────────────────────────────────────────────────
    def fmt_outcomes(lst: list[dict]) -> str:
        lines = []
        for o in lst:
            m = o.get("measure", "").strip()
            tf = o.get("timeFrame", "").strip()
            lines.append(f"• {m}" + (f"  [{tf}]" if tf else ""))
        return "\n".join(lines)

    primary_outcomes   = fmt_outcomes(outcomes_mod.get("primaryOutcomes",   []))
    secondary_outcomes = fmt_outcomes(outcomes_mod.get("secondaryOutcomes", []))

    # ── Eligibility ───────────────────────────────────────────────────────
    elig_text = elig_mod.get("eligibilityCriteria", "")
    inclusion, exclusion = parse_eligibility(elig_text)
    min_age = elig_mod.get("minimumAge", "")
    max_age = elig_mod.get("maximumAge", "")
    sex     = elig_mod.get("sex", "")

    # ── Cancer type ───────────────────────────────────────────────────────
    cancer_types    = classify_cancer_types(conditions_full, title)
    cancer_type_str = " | ".join(cancer_types)

    # ── Global PI ─────────────────────────────────────────────────────────
    global_pi, global_pi_affil = extract_pi(study)

    # ── Site rows (one per qualifying SoCal location) ─────────────────────
    site_rows: list[dict] = []
    for loc in contacts_mod.get("locations", []):
        state    = loc.get("state",    "")
        city     = loc.get("city",     "")
        facility = loc.get("facility", "")

        if state and state not in ("California", "CA"):
            continue
        if not is_socal_site(city, facility):
            continue
        institution = normalize_facility(facility)
        if institution not in TARGET_INSTITUTIONS:
            continue

        site_pi, site_email, site_phone = extract_site_contact(loc)
        pi_name  = site_pi    if site_pi    else global_pi
        pi_affil = facility   if site_pi    else global_pi_affil
        pi_email = site_email  # email is site-specific; no global fallback intentionally

        site_rows.append({
            "NCT ID":         nct_id,
            "Trial title":    title,
            "Phase":          phase,
            "Status":         status,
            "Cancer type":    cancer_type_str,
            "Institution":    institution,
            "PI name":        pi_name,
            "PI email":       pi_email,
            "PI phone":       site_phone,
            "PI affiliation": pi_affil,
            "Site city":      city,
            "Lead sponsor":   lead_sponsor,
        })

    if not site_rows:
        return None, []

    # ── Study record (one per trial) ──────────────────────────────────────
    study_record: dict = {
        # ── Identification
        "NCT ID":                  nct_id,
        "Study title":             title,
        "Other study IDs":         sec_ids,
        "ClinicalTrials URL":      f"https://clinicaltrials.gov/study/{nct_id}",
        # ── Classification
        "Phase":                   phase,
        "Status":                  status,
        "Cancer type":             cancer_type_str,
        "Conditions":              conditions_full,
        # ── Sponsor / responsible party
        "Lead sponsor":            lead_sponsor,
        "Additional sponsors":     additional_sponsors,
        "Information provided by": info_provided_by,
        # ── Intervention
        "Intervention(s)":         intervention_str,
        # ── Design + enrollment
        "Study design":            design_str,
        "Enrollment (estimated)":  enroll_est,
        "Enrollment (actual)":     enroll_act,
        # ── Dates
        "Start date":              start_date,
        "Primary completion":      end_date,
        "Last update posted":      last_update,
        "Study first posted":      first_post_date,
        "Last pulled by":          PULLED_BY,
        # ── Narrative
        "Brief summary":           brief_summary,
        # ── Eligibility (kept separate — stored in study_record for Sheet 5)
        "Min age":                 min_age,
        "Max age":                 max_age,
        "Sex":                     sex,
        "Inclusion criteria":      inclusion,
        "Exclusion criteria":      exclusion,
        # ── Raw fields needed by classify_trial (not written to all sheets)
        "_interventions_raw":      intervention_str,
        "_inclusion_raw":          inclusion,
        "_exclusion_raw":          exclusion,
        # ── Outcomes (stored in study_record for Sheet 6)
        "Primary outcomes":        primary_outcomes,
        "Secondary outcomes":      secondary_outcomes,
    }

    return study_record, site_rows


# ── API fetch (identical to master pipeline) ──────────────────────────────────
def _fetch_condition_pages(condition: str) -> list[dict]:
    params: dict = {
        "query.cond":           condition,
        "filter.geo":           GEO_FILTER,
        "filter.overallStatus": "|".join(STATUSES),
        "filter.advanced":      f"AREA[StudyType]{STUDY_TYPE}",
        "pageSize":             PAGE_SIZE,
        "format":               "json",
    }
    studies: list[dict] = []
    nxt: str | None = None
    while True:
        if nxt:
            params["pageToken"] = nxt
        resp = requests.get(BASE_URL, params=params, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        studies.extend(data.get("studies", []))
        nxt = data.get("nextPageToken")
        if not nxt:
            break
        time.sleep(0.05)
    return studies


def fetch_all() -> list[dict]:
    seen: set[str]  = set()
    all_: list[dict] = []
    lock = threading.Lock()

    def worker(cond: str) -> tuple[str, int]:
        try:
            batch = _fetch_condition_pages(cond)
        except Exception as exc:
            return cond, -1
        added = 0
        with lock:
            for s in batch:
                nct = (s.get("protocolSection", {})
                        .get("identificationModule", {})
                        .get("nctId", ""))
                if nct and nct not in seen:
                    seen.add(nct); all_.append(s); added += 1
        return cond, added

    print(f"Fetching {len(CONDITIONS)} conditions in parallel…\n")
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        for cond, n in pool.map(worker, CONDITIONS):
            status_str = f"+{n}" if n >= 0 else "FAILED"
            print(f"  {cond:<32} {status_str}")

    # Client-side safety filter
    before = len(all_)
    all_ = [
        s for s in all_
        if (s.get("protocolSection", {})
              .get("designModule", {})
              .get("studyType", "")
              .upper() == STUDY_TYPE)
        and (s.get("protocolSection", {})
               .get("statusModule", {})
               .get("overallStatus", "")
             in STATUSES)
    ]
    dropped = before - len(all_)
    if dropped:
        print(f"\n  [filter] dropped {dropped} non-interventional/wrong-status records")
    print(f"\n  Total unique studies: {len(all_)}")
    return all_


# ── Excel styling ─────────────────────────────────────────────────────────────
_C = {
    "navy":        "1F4E79",
    "blue":        "2E75B6",
    "red":         "C00000",
    "green":       "375623",
    "teal":        "1F6B75",
    "gold":        "7B5C00",
    "white":       "FFFFFF",
    "link":        "0563C1",
    "alt":         "DCE6F1",
    "alt2":        "FFF2CC",
    "dark_gray":   "404040",
    "light_gray":  "F2F2F2",
    "conf_high":   "C6EFCE",   # light green fill
    "conf_medium": "FFEB9C",   # light yellow fill
    "conf_low":    "FFCC99",   # light orange fill
    # Institution map cell colors
    "map_zero":    "F2F2F2",
    "map_low":     "D6E4F0",
    "map_mid":     "2E6DA4",
    "map_high":    "1B3A6B",
    # Section header colors (from taxonomy)
    "sec_nmibc":       "1B7A4E",
    "sec_mibc":        "1B3A6B",
    "sec_localadv":    "8B6914",
    "sec_metastatic":  "8B3A2F",
}

_H_FONT  = Font(bold=True, color=_C["white"], name="Arial", size=10)
_B_FONT  = Font(name="Arial", size=10)
_L_FONT  = Font(color=_C["link"], underline="single", name="Arial", size=10)
_H_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
_W_ALIGN = Alignment(wrap_text=True, vertical="top")


def _style_ws(ws, tab_color: str, url_col: str | None = None,
              wrap_cols: list[str] | None = None,
              freeze: str = "A2", row_h: int = 15) -> None:
    """Apply header + body styling. wrap_cols: list of header names to text-wrap."""
    wrap_set   = set(wrap_cols or [])
    url_col_no = None

    for cell in ws[1]:
        cell.font      = _H_FONT
        cell.fill      = PatternFill("solid", fgColor=tab_color)
        cell.alignment = _H_ALIGN
        if cell.value == url_col:
            url_col_no = cell.column

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = freeze
    ws.sheet_properties.tabColor = tab_color

    # Body rows
    for ri, row in enumerate(ws.iter_rows(min_row=2), start=2):
        alt = (ri % 2 == 0)
        for cell in row:
            is_link = url_col_no and cell.column == url_col_no
            header  = ws.cell(1, cell.column).value or ""
            wrap    = header in wrap_set

            cell.font = _L_FONT if is_link else _B_FONT
            if alt:
                cell.fill = PatternFill("solid", fgColor=_C["alt"])
            cell.alignment = _W_ALIGN if wrap else Alignment(vertical="top")

            if is_link and cell.value:
                cell.hyperlink = str(cell.value)

        ws.row_dimensions[ri].height = row_h

    # Column widths
    for ci, col in enumerate(ws.columns, 1):
        hdr  = ws.cell(1, ci).value or ""
        mx   = max((len(str(c.value or "")) for c in col), default=10)
        if hdr in (wrap_set or set()):
            ws.column_dimensions[get_column_letter(ci)].width = 70
        else:
            ws.column_dimensions[get_column_letter(ci)].width = min(mx + 3, 55)


def _df_to_ws(ws, df: pd.DataFrame) -> None:
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append([None if (isinstance(v, float) and pd.isna(v)) else v for v in row])


# ── Helper: phase sort key ────────────────────────────────────────────────────
def _phase_sort_key(phase: str) -> int:
    """Higher = later phase. Used to sort trials within a disease setting."""
    p = (phase or "").upper()
    if "3" in p:
        return 4
    if "2" in p and "3" in p:
        return 3
    if "2" in p:
        return 3
    if "1" in p and "2" in p:
        return 2
    if "1" in p:
        return 1
    return 0


# ── Sheet 1: Trial Finder ─────────────────────────────────────────────────────
def _build_trial_finder(ws, study_records: list[dict], site_rows: list[dict],
                        cancer_type: str = "Bladder/Urothelial",
                        cancer_type_filters: list[str] | None = None) -> tuple[int, int]:
    """
    Build the physician-facing Trial Finder sheet.
    cancer_type: the NCCN canonical name used for disease settings/colors.
    cancer_type_filters: list of "Cancer type" strings to include (defaults to [cancer_type]).
    Returns (n_settings_written, n_trial_entries_written).
    """
    if cancer_type_filters is None:
        cancer_type_filters = [cancer_type]

    ws.sheet_properties.tabColor = _C["navy"]
    ws.freeze_panes = "A2"

    # Section color map — load from taxonomy; fall back to built-in palette
    _tax_colors = get_section_colors(cancer_type)  # {section: hex}
    _palette = ["1F4E79","2E75B6","375623","1F6B75","7B5C00","C00000","404040","1B3A6B"]
    _pal_idx = [0]
    def _sec_color(section: str) -> str:
        if section in _tax_colors:
            return _tax_colors[section]
        c = _palette[_pal_idx[0] % len(_palette)]
        _pal_idx[0] += 1
        return c
    section_color_map: dict[str, str] = {}

    # Column definitions
    col_headers = ["NCT ID", "Trial Title", "Phase", "Treatment", "Available At",
                   "Top PI(s)", "Confidence", "Link"]
    col_widths  = [14, 55, 8, 28, 45, 32, 11, 8]

    # Set column widths up front
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Build lookup dicts from site_rows
    # {nct_id: sorted list of unique institutions}
    nct_to_institutions: dict[str, list[str]] = defaultdict(set)
    # {nct_id: list of unique PI names (excluding "Not listed")}
    nct_to_pis: dict[str, list[str]] = defaultdict(list)
    # {nct_id: dict of canon_key -> email}  — first email seen per PI wins
    nct_to_pi_email: dict[str, dict[str, str]] = defaultdict(dict)
    _seen_pi_keys: dict[str, set] = defaultdict(set)

    for sr in site_rows:
        nid = sr["NCT ID"]
        nct_to_institutions[nid].add(sr["Institution"])
        pi = sr.get("PI name", "")
        if pi and pi != "Not listed":
            canon = canonical_pi(pi)
            if canon and canon not in _seen_pi_keys[nid]:
                _seen_pi_keys[nid].add(canon)
                nct_to_pis[nid].append(pi)
            # Always capture email even if we already saw this PI (first non-empty wins)
            email = sr.get("PI email", "")
            if canon and email and canon not in nct_to_pi_email[nid]:
                nct_to_pi_email[nid][canon] = email

    # Convert sets to sorted lists
    nct_to_inst_sorted = {k: sorted(v) for k, v in nct_to_institutions.items()}

    # Build index: nct_id -> study_record
    nct_to_record = {r["NCT ID"]: r for r in study_records}

    # Get ordered disease settings for this cancer type
    ds_list = get_disease_settings_in_order(cancer_type)

    # Font/fill helpers (defined locally for clarity)
    white_bold = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    white_norm = Font(color="FFFFFF", name="Arial", size=10)
    dark_bold  = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    col_hdr_f  = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    body_font  = Font(name="Arial", size=9)
    link_font  = Font(color=_C["link"], underline="single", name="Arial", size=9)
    center_al  = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left_al    = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    left_mid   = Alignment(horizontal="left",   vertical="center", wrap_text=False)

    conf_fills = {
        "HIGH":   PatternFill("solid", fgColor=_C["conf_high"]),
        "MEDIUM": PatternFill("solid", fgColor=_C["conf_medium"]),
        "LOW":    PatternFill("solid", fgColor=_C["conf_low"]),
    }
    conf_fonts = {
        "HIGH":   Font(name="Arial", size=9, color="276221"),
        "MEDIUM": Font(name="Arial", size=9, color="7A5C00"),
        "LOW":    Font(name="Arial", size=9, color="8B3000"),
    }

    current_row  = 1
    last_section = None
    n_settings   = 0
    n_entries    = 0

    for ds in ds_list:
        label   = ds["label"]
        short   = ds["short"]
        section = ds["section"]

        # Find trials matching this disease setting
        matching = [
            r for r in study_records
            if any(ct in r.get("Cancer type", "") for ct in cancer_type_filters)
            and label in r.get("Disease setting (all)", "")
        ]

        if not matching:
            continue

        n_settings += 1
        sec_color = _sec_color(section)

        # Lighten section header slightly for category rows
        # (use same color but treat section vs category differently)
        # Section header: only write when section changes
        if section != last_section:
            last_section = section
            # Merge A:H
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row,   end_column=8
            )
            cell = ws.cell(current_row, 1, section.upper())
            cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
            cell.fill      = PatternFill("solid", fgColor=sec_color)
            cell.alignment = Alignment(horizontal="left", vertical="center",
                                       indent=1, wrap_text=False)
            ws.row_dimensions[current_row].height = 22
            current_row += 1

        # Category header row
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row,   end_column=8
        )
        cat_label_text = f"{label}  —  {len(matching)} trial{'s' if len(matching) != 1 else ''}"
        cat_cell = ws.cell(current_row, 1, cat_label_text)
        # Slightly lighter shade: add 30 to each hex component, clamped
        r_int = min(int(sec_color[0:2], 16) + 30, 255)
        g_int = min(int(sec_color[2:4], 16) + 30, 255)
        b_int = min(int(sec_color[4:6], 16) + 30, 255)
        lighter = f"{r_int:02X}{g_int:02X}{b_int:02X}"
        cat_cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cat_cell.fill      = PatternFill("solid", fgColor=lighter)
        cat_cell.alignment = Alignment(horizontal="left", vertical="center",
                                       indent=2, wrap_text=False)
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        # Column header row
        col_hdr_fill = PatternFill("solid", fgColor=_C["dark_gray"])
        for ci, hdr in enumerate(col_headers, 1):
            c = ws.cell(current_row, ci, hdr)
            c.font      = col_hdr_f
            c.fill      = col_hdr_fill
            c.alignment = center_al
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # Sort trials: phase descending, then institution count descending
        matching_sorted = sorted(
            matching,
            key=lambda r: (
                -_phase_sort_key(r.get("Phase", "")),
                -len(nct_to_inst_sorted.get(r["NCT ID"], [])),
            )
        )

        for row_idx, rec in enumerate(matching_sorted):
            nid        = rec["NCT ID"]
            trial_title = (rec.get("Study title") or "")[:75]
            phase_val  = rec.get("Phase", "N/A")
            treatment  = rec.get("Treatment modality", "")
            avail_at   = ", ".join(nct_to_inst_sorted.get(nid, []))
            # Build "PI Name (email)" where CT.gov provides a contact email
            _pi_email_map = nct_to_pi_email.get(nid, {})
            _pi_parts = []
            for _pn in nct_to_pis.get(nid, [])[:3]:
                _canon = canonical_pi(_pn)
                _em    = _pi_email_map.get(_canon, "")
                _pi_parts.append(f"{_pn} ({_em})" if _em else _pn)
            top_pis = "; ".join(_pi_parts)
            confidence = rec.get("Classification confidence", "")
            ct_url     = f"https://clinicaltrials.gov/study/{nid}"

            is_alt = (row_idx % 2 == 1)
            row_fill = PatternFill("solid", fgColor="F7F7F7") if is_alt else None

            def _wc(col: int, val, font=None, fill=None, align=None, hyperlink=None):
                c = ws.cell(current_row, col, val)
                c.font = font or body_font
                if fill:
                    c.fill = fill
                elif row_fill and col not in (7,):   # confidence col has own fill
                    c.fill = row_fill
                c.alignment = align or left_al
                if hyperlink:
                    c.hyperlink = hyperlink

            row_bg = row_fill

            # Col 1: NCT ID (hyperlinked)
            c1 = ws.cell(current_row, 1, nid)
            c1.font      = link_font
            c1.hyperlink = ct_url
            c1.alignment = left_mid
            if row_bg:
                c1.fill = row_bg

            # Col 2: Title
            c2 = ws.cell(current_row, 2, trial_title)
            c2.font      = body_font
            c2.alignment = left_al
            if row_bg:
                c2.fill = row_bg

            # Col 3: Phase
            c3 = ws.cell(current_row, 3, phase_val)
            c3.font      = body_font
            c3.alignment = center_al
            if row_bg:
                c3.fill = row_bg

            # Col 4: Treatment
            c4 = ws.cell(current_row, 4, treatment)
            c4.font      = body_font
            c4.alignment = left_al
            if row_bg:
                c4.fill = row_bg

            # Col 5: Available At
            c5 = ws.cell(current_row, 5, avail_at)
            c5.font      = body_font
            c5.alignment = left_al
            if row_bg:
                c5.fill = row_bg

            # Col 6: Top PIs
            c6 = ws.cell(current_row, 6, top_pis)
            c6.font      = body_font
            c6.alignment = left_al
            if row_bg:
                c6.fill = row_bg

            # Col 7: Confidence (with color fill)
            c7 = ws.cell(current_row, 7, confidence)
            c7.font      = conf_fonts.get(confidence, body_font)
            c7.fill      = conf_fills.get(confidence, PatternFill())
            c7.alignment = center_al

            # Col 8: Link arrow
            c8 = ws.cell(current_row, 8, "→ CT.gov")
            c8.font      = link_font
            c8.hyperlink = ct_url
            c8.alignment = center_al
            if row_bg:
                c8.fill = row_bg

            ws.row_dimensions[current_row].height = 16
            current_row += 1
            n_entries += 1

        # Spacer row after each category
        ws.row_dimensions[current_row].height = 6
        current_row += 1

    return n_settings, n_entries


# ── Sheet 2: Institution Map ───────────────────────────────────────────────────
def _build_institution_map(ws, study_records: list[dict], site_rows: list[dict],
                           max_institutions: int = 18) -> tuple[int, int]:
    """
    Build the disease setting × institution matrix.
    Returns (n_settings, n_institutions).
    """
    ws.sheet_properties.tabColor = _C["teal"]

    # Get ordered disease settings (bladder only for now)
    ds_list = get_disease_settings_in_order("Bladder/Urothelial")

    # Filter to settings that have at least 1 matching trial
    bladder_records = [r for r in study_records if "Bladder/Urothelial" in r.get("Cancer type", "")]

    # Count trials per institution (for sorting institutions)
    inst_trial_counts: dict[str, set] = defaultdict(set)
    for sr in site_rows:
        nid  = sr["NCT ID"]
        inst = sr["Institution"]
        # Only count bladder trials in the institution map
        rec  = next((r for r in bladder_records if r["NCT ID"] == nid), None)
        if rec:
            inst_trial_counts[inst].add(nid)

    # Sort institutions by total trial count desc, then alphabetically
    institutions_sorted = sorted(
        inst_trial_counts.keys(),
        key=lambda i: (-len(inst_trial_counts[i]), i)
    )[:max_institutions]

    if not institutions_sorted:
        # No bladder data — write placeholder
        ws.cell(1, 1, "No Bladder/Urothelial trials found for Institution Map.")
        return 0, 0

    # Build {(label, institution): set of NCT IDs}
    cell_data: dict[tuple[str, str], set] = defaultdict(set)
    for rec in bladder_records:
        nid = rec["NCT ID"]
        ds_all = rec.get("Disease setting (all)", "")
        for ds in ds_list:
            if ds["label"] in ds_all:
                for sr in site_rows:
                    if sr["NCT ID"] == nid and sr["Institution"] in institutions_sorted:
                        cell_data[(ds["label"], sr["Institution"])].add(nid)

    # ── Header row ────────────────────────────────────────────────────────
    navy_fill = PatternFill("solid", fgColor=_C["navy"])
    hdr_font  = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    hdr_al    = Alignment(horizontal="center", vertical="center",
                          wrap_text=True, text_rotation=45)
    left_al   = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_al = Alignment(horizontal="center", vertical="center")

    ws.cell(1, 1, "Disease Setting").font      = hdr_font
    ws.cell(1, 1).fill                          = navy_fill
    ws.cell(1, 1).alignment                     = left_al
    ws.column_dimensions["A"].width             = 32
    ws.row_dimensions[1].height                 = 22

    for ci, inst in enumerate(institutions_sorted, 2):
        c = ws.cell(1, ci, inst)
        c.font      = hdr_font
        c.fill      = navy_fill
        c.alignment = hdr_al
        ws.column_dimensions[get_column_letter(ci)].width = 9

    ws.freeze_panes = "B2"

    # Section color map for row shading
    section_color_map = {
        "NMIBC":           "E8F5EE",
        "MIBC":            "E8EDF5",
        "Locally Advanced":"F5F0E0",
        "Metastatic":      "F5E8E8",
    }

    # Cell fill logic by count
    def _count_fill(n: int) -> tuple[PatternFill, Font]:
        if n == 0:
            return PatternFill("solid", fgColor=_C["map_zero"]), Font(name="Arial", size=9, color="AAAAAA")
        elif n <= 2:
            return PatternFill("solid", fgColor=_C["map_low"]),  Font(name="Arial", size=9, color="1B3A6B")
        elif n <= 5:
            return PatternFill("solid", fgColor=_C["map_mid"]),  Font(bold=True, name="Arial", size=9, color="FFFFFF")
        else:
            return PatternFill("solid", fgColor=_C["map_high"]), Font(bold=True, name="Arial", size=9, color="FFFFFF")

    body_font_map  = Font(name="Arial", size=9)
    n_settings_written = 0

    for ri, ds in enumerate(ds_list, 2):
        label   = ds["label"]
        short   = ds["short"]
        section = ds["section"]

        # Check if this setting has any data
        has_data = any(
            cell_data.get((label, inst), set())
            for inst in institutions_sorted
        )
        if not has_data:
            continue

        n_settings_written += 1
        row_bg_hex = section_color_map.get(section, "FFFFFF")
        row_fill   = PatternFill("solid", fgColor=row_bg_hex)

        # Col A: disease setting label (short)
        ca = ws.cell(ri, 1, short)
        ca.font      = body_font_map
        ca.fill      = row_fill
        ca.alignment = left_al

        # Cols B+: counts
        for ci, inst in enumerate(institutions_sorted, 2):
            count = len(cell_data.get((label, inst), set()))
            disp  = str(count) if count > 0 else ""
            c_fill, c_font = _count_fill(count)
            c = ws.cell(ri, ci, disp)
            c.font      = c_font
            c.fill      = c_fill
            c.alignment = center_al

        ws.row_dimensions[ri].height = 16

    return n_settings_written, len(institutions_sorted)


# ── Build Excel (8 sheets) ────────────────────────────────────────────────────
def build_excel(study_records: list[dict], site_rows: list[dict],
                run_ts: str, out_path: Path) -> None:

    df_studies = pd.DataFrame(study_records)
    df_sites   = pd.DataFrame(site_rows)

    # ── Sheet 3: Trial Details ─────────────────────────────────────────────
    cols_detail = [
        "NCT ID", "Study title", "Phase", "Status", "Cancer type",
        "Disease setting (primary)", "Disease setting (all)",
        "Classification confidence", "BCG status", "Cisplatin status",
        "Treatment modality", "Delivery",
        "Conditions", "Lead sponsor", "Additional sponsors",
        "Information provided by", "Intervention(s)", "Study design",
        "Enrollment (estimated)", "Enrollment (actual)", "Other study IDs",
        "Start date", "Primary completion", "Last update posted",
        "Last pulled by", "Brief summary", "ClinicalTrials URL",
    ]
    df_detail = df_studies[[c for c in cols_detail if c in df_studies.columns]].copy()

    # ── Sheet 4: Sites × PI ────────────────────────────────────────────────
    cols_sites = [
        "NCT ID", "Trial title", "Phase", "Status", "Cancer type",
        "Disease setting (primary)", "Classification confidence",
        "Treatment modality", "Delivery",
        "Institution", "PI name", "PI email", "PI phone", "PI affiliation", "Site city", "Lead sponsor",
    ]
    df_sites_out = df_sites[[c for c in cols_sites if c in df_sites.columns]].copy()
    df_sites_out = df_sites_out.drop_duplicates(["NCT ID", "Institution"])

    # ── Sheet 5: Eligibility ───────────────────────────────────────────────
    cols_elig = [
        "NCT ID", "Study title", "Cancer type",
        "Disease setting (primary)", "Phase",
        "Min age", "Max age", "Sex",
        "Inclusion criteria", "Exclusion criteria",
    ]
    df_elig = df_studies[[c for c in cols_elig if c in df_studies.columns]].copy()

    # ── Sheet 6: Outcomes ──────────────────────────────────────────────────
    cols_out = [
        "NCT ID", "Study title", "Cancer type",
        "Disease setting (primary)", "Phase",
        "Primary outcomes", "Secondary outcomes",
    ]
    df_outcomes = df_studies[[c for c in cols_out if c in df_studies.columns]].copy()

    # ── Sheet 7: PI Summary ────────────────────────────────────────────────
    df_s = df_sites_out.copy()
    df_s["canon"] = df_s["PI name"].apply(canonical_pi)
    df_valid = df_s[(df_s["PI name"] != "Not listed") & (df_s["canon"] != "")]

    if not df_valid.empty:
        pi_inst = (
            df_valid
            .groupby(["Institution", "canon"], sort=False)
            .agg(
                PI_Name =("PI name",   lambda x: x.value_counts().index[0]),
                Trials  =("NCT ID",    "nunique"),
                Types   =("Cancer type",lambda x: " | ".join(sorted({
                              t for v in x for t in v.split(" | ")
                          }))),
            )
            .reset_index()
            .sort_values(["Institution", "Trials"], ascending=[True, False])
            [["Institution", "PI_Name", "Trials", "Types"]]
            .rename(columns={"PI_Name": "PI Name", "Types": "Cancer types"})
        )
    else:
        pi_inst = pd.DataFrame(columns=["Institution", "PI Name", "Trials", "Cancer types"])

    # ── Assemble workbook ──────────────────────────────────────────────────
    wb       = Workbook()
    ws_tf    = wb.active;                      ws_tf.title    = "Trial Finder"
    ws_map   = wb.create_sheet("Institution Map")
    ws3      = wb.create_sheet("Trial Details")
    ws4      = wb.create_sheet("Sites × PI")
    ws5      = wb.create_sheet("Eligibility")
    ws6      = wb.create_sheet("Outcomes")
    ws7      = wb.create_sheet("PI Summary")
    ws_meta  = wb.create_sheet("Run Metadata")

    # Sheet 1: Trial Finder (custom builder)
    tf_settings, tf_entries = _build_trial_finder(ws_tf, study_records, site_rows)

    # Sheet 2: Institution Map (custom builder)
    map_settings, map_insts = _build_institution_map(ws_map, study_records, site_rows)

    # Sheets 3–7: data frames
    _df_to_ws(ws3, df_detail)
    _df_to_ws(ws4, df_sites_out)
    _df_to_ws(ws5, df_elig)
    _df_to_ws(ws6, df_outcomes)
    _df_to_ws(ws7, pi_inst)

    _style_ws(ws3, _C["navy"],  url_col="ClinicalTrials URL",
              wrap_cols=["Brief summary", "Intervention(s)", "Study design"],
              row_h=30)
    _style_ws(ws4, _C["blue"],  row_h=16)
    _style_ws(ws5, _C["red"],
              wrap_cols=["Inclusion criteria", "Exclusion criteria"],
              row_h=60)
    _style_ws(ws6, _C["green"],
              wrap_cols=["Primary outcomes", "Secondary outcomes"],
              row_h=45)
    _style_ws(ws7, _C["teal"],  row_h=16)

    # ── Run Metadata sheet ────────────────────────────────────────────────
    ws_meta.sheet_properties.tabColor = _C["gold"]
    nccn_stamp = get_nccn_stamp("Bladder/Urothelial")

    meta_rows = [
        ("Field",                "Value"),
        ("Script",               f"gu_oncology_extended_v{VERSION}.py"),
        ("Script version",       VERSION),
        ("Run date",             run_ts.split("_")[0]),
        ("Run time",             run_ts.split("_")[1]),
        ("Pulled by",            PULLED_BY),
        ("Data source",          "ClinicalTrials.gov API v2"),
        ("API base URL",         BASE_URL),
        ("",                     ""),
        ("── NCCN Classification ──", ""),
        ("NCCN taxonomy",        nccn_stamp),
        ("Classification scope", "Bladder/Urothelial (Prostate, Renal: pending)"),
        ("",                     ""),
        ("── Filters ──",        ""),
        ("Study type",           STUDY_TYPE),
        ("Status",               ", ".join(STATUSES)),
        ("Geography",            GEO_FILTER),
        ("Date cutoff",          "NONE — all active trials"),
        ("Conditions queried",   str(len(CONDITIONS))),
        *[(f"  Condition {i+1}", c) for i, c in enumerate(CONDITIONS)],
        ("Target institutions",  str(len(TARGET_INSTITUTIONS))),
        ("",                     ""),
        ("── Output ──",         ""),
        ("Unique trials",        str(len(df_studies))),
        ("Site × institution rows", str(len(df_sites_out))),
        ("Institutions found",   str(df_sites_out["Institution"].nunique()
                                     if not df_sites_out.empty else 0)),
        ("Output file",          out_path.name),
    ]

    hdr_font  = Font(bold=True, color=_C["white"], name="Arial", size=10)
    hdr_fill  = PatternFill("solid", fgColor=_C["gold"])
    body_font = Font(name="Arial", size=10)
    key_font  = Font(bold=True, name="Arial", size=10)

    for i, (k, v) in enumerate(meta_rows, start=1):
        ws_meta.cell(i, 1, k)
        ws_meta.cell(i, 2, v)
        if i == 1:
            ws_meta.cell(i, 1).font = hdr_font
            ws_meta.cell(i, 1).fill = hdr_fill
            ws_meta.cell(i, 2).font = hdr_font
            ws_meta.cell(i, 2).fill = hdr_fill
        elif k.startswith("──") or k == "":
            ws_meta.cell(i, 1).font = Font(bold=True, color="888888",
                                           name="Arial", size=9, italic=True)
        else:
            ws_meta.cell(i, 1).font = key_font
            ws_meta.cell(i, 2).font = body_font

    ws_meta.column_dimensions["A"].width = 28
    ws_meta.column_dimensions["B"].width = 55

    wb.save(out_path)

    print(f"\n  Excel saved → {out_path.name}")
    print(f"    Sheet 1 — Trial Finder        : {tf_settings:>3} disease settings, {tf_entries:>4} trial entries")
    print(f"    Sheet 2 — Institution Map     : {map_settings:>3} settings × {map_insts:>2} institutions")
    print(f"    Sheet 3 — Trial Details       : {len(df_detail):>4} trials")
    print(f"    Sheet 4 — Sites × PI          : {len(df_sites_out):>4} rows")
    print(f"    Sheet 5 — Eligibility         : {len(df_elig):>4} trials")
    print(f"    Sheet 6 — Outcomes            : {len(df_outcomes):>4} trials")
    print(f"    Sheet 7 — PI Summary          : {len(pi_inst):>4} PI-institution pairs")
    print(f"    Sheet 8 — Run Metadata")



# ── Find previous CSV for delta ───────────────────────────────────────────────
def _find_prev_csv(out_root: Path, current_ts: str) -> "Path | None":
    """Return the most recent all_trials CSV from a prior run, or None."""
    candidates = []
    for d in out_root.iterdir():
        if not d.is_dir():
            continue
        if d.name >= current_ts:
            continue
        for csv_file in d.glob("all_trials_*.csv"):
            candidates.append(csv_file)
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.parent.name)


# ── Per-cancer-type Excel (4 sheets) ─────────────────────────────────────────
def build_cancer_type_excel(
    cancer_type: str,
    cancer_type_filters: list[str],
    study_records: list[dict],
    site_rows: list[dict],
    run_ts: str,
    out_path: "Path",
) -> None:
    """
    Build a 4-sheet workbook for one cancer type group.
    Sheets: Trial Finder · Sites × PI · Eligibility · Outcomes
    """
    # Filter to this cancer type group
    ct_recs  = [r for r in study_records
                if any(ct in r.get("Cancer type", "") for ct in cancer_type_filters)]
    ct_sites = [r for r in site_rows
                if any(ct in r.get("Cancer type", "") for ct in cancer_type_filters)]

    if not ct_recs:
        print(f"  [SKIP] No trials found for {cancer_type_filters}")
        return

    df_recs  = pd.DataFrame(ct_recs)
    df_sites = pd.DataFrame(ct_sites)

    # ── Sheet 2: Sites × PI
    cols_sites = [
        "NCT ID", "Trial title", "Phase", "Status", "Cancer type",
        "Disease setting (primary)", "Classification confidence",
        "Treatment modality", "Delivery",
        "Institution", "PI name", "PI email", "PI phone", "PI affiliation",
        "Site city", "Lead sponsor",
    ]
    df_sites_out = (df_sites[[c for c in cols_sites if c in df_sites.columns]]
                    .copy()
                    .drop_duplicates(["NCT ID", "Institution"]))

    # ── Sheet 3: Eligibility
    cols_elig = [
        "NCT ID", "Study title", "Cancer type",
        "Disease setting (primary)", "Phase",
        "Min age", "Max age", "Sex",
        "Inclusion criteria", "Exclusion criteria",
    ]
    df_elig = df_recs[[c for c in cols_elig if c in df_recs.columns]].copy()

    # ── Sheet 4: Outcomes
    cols_out = [
        "NCT ID", "Study title", "Cancer type",
        "Disease setting (primary)", "Phase",
        "Primary outcomes", "Secondary outcomes",
    ]
    df_outcomes = df_recs[[c for c in cols_out if c in df_recs.columns]].copy()

    # ── Assemble workbook
    wb     = Workbook()
    ws_tf  = wb.active;                  ws_tf.title  = "Trial Finder"
    ws_sp  = wb.create_sheet("Sites × PI")
    ws_el  = wb.create_sheet("Eligibility")
    ws_out = wb.create_sheet("Outcomes")

    # Sheet 1: Trial Finder (generalised by cancer_type)
    tf_settings, tf_entries = _build_trial_finder(
        ws_tf, ct_recs, ct_sites,
        cancer_type=cancer_type,
        cancer_type_filters=cancer_type_filters,
    )

    # Sheets 2–4: data frames
    _df_to_ws(ws_sp,  df_sites_out)
    _df_to_ws(ws_el,  df_elig)
    _df_to_ws(ws_out, df_outcomes)

    _style_ws(ws_sp,  _C["blue"],  row_h=16)
    _style_ws(ws_el,  _C["red"],
              wrap_cols=["Inclusion criteria", "Exclusion criteria"], row_h=60)
    _style_ws(ws_out, _C["green"],
              wrap_cols=["Primary outcomes", "Secondary outcomes"],   row_h=45)

    wb.save(out_path)

    label = " / ".join(cancer_type_filters)
    print(f"  {out_path.name}")
    print(f"    Trial Finder : {tf_settings:>2} settings, {tf_entries:>3} entries")
    print(f"    Sites × PI   : {len(df_sites_out):>3} rows")
    print(f"    Eligibility  : {len(df_elig):>3} trials")
    print(f"    Outcomes     : {len(df_outcomes):>3} trials")


# ── Updates / analytics Excel (4 sheets) ─────────────────────────────────────
def build_updates_excel(
    study_records: list[dict],
    site_rows: list[dict],
    run_ts: str,
    prev_csv: "Path | None",
    out_path: "Path",
    recall_result: dict | None = None,
) -> None:
    """
    Build the updates/analytics workbook.
    Sheets: Changes · Stats by Hospital · New Trials by Window · Metadata
    """
    from datetime import datetime, timedelta

    run_date = run_ts.split("_")[0]
    today    = datetime.strptime(run_date, "%Y-%m-%d")
    six_mo   = (today - timedelta(days=183)).strftime("%Y-%m-%d")
    twelve_mo= (today - timedelta(days=365)).strftime("%Y-%m-%d")

    df_cur = pd.DataFrame(study_records)
    df_sites = pd.DataFrame(site_rows) if site_rows else pd.DataFrame()

    wb     = Workbook()
    ws_ch  = wb.active;                   ws_ch.title  = "Changes since last pull"
    ws_st  = wb.create_sheet("Stats by Hospital")
    ws_nw  = wb.create_sheet("New Trials by Window")
    ws_me  = wb.create_sheet("Metadata")

    gold_fill  = PatternFill("solid", fgColor=_C["gold"])
    navy_fill  = PatternFill("solid", fgColor=_C["navy"])
    hdr_font   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    body_font  = Font(name="Arial", size=10)
    key_font   = Font(bold=True, name="Arial", size=10)
    center_al  = Alignment(horizontal="center", vertical="center")
    left_al    = Alignment(horizontal="left",   vertical="top", wrap_text=True)

    # ── Sheet 1: Changes ──────────────────────────────────────────────────────
    ws_ch.sheet_properties.tabColor = _C["red"]
    cur_ncts = set(df_cur["NCT ID"].tolist()) if "NCT ID" in df_cur.columns else set()

    if prev_csv is None or not prev_csv.exists():
        # First run — no prior pull
        ws_ch.append(["Status", "NCT ID", "Trial Title", "Cancer Type",
                       "Phase", "Institution(s)", "Note"])
        for cell in ws_ch[1]:
            cell.font = hdr_font; cell.fill = navy_fill; cell.alignment = center_al
        ws_ch.append(["BASELINE", "", "First run — no prior pull available for comparison.",
                       "", "", "", f"Baseline established {run_date}. All {len(cur_ncts)} trials are new."])
        ws_ch.column_dimensions["A"].width = 12
        ws_ch.column_dimensions["B"].width = 14
        ws_ch.column_dimensions["C"].width = 60
        ws_ch.column_dimensions["G"].width = 50
    else:
        df_prev = pd.read_csv(prev_csv, dtype=str).fillna("")
        prev_ncts = set(df_prev["NCT ID"].tolist()) if "NCT ID" in df_prev.columns else set()
        prev_date = prev_csv.parent.name.split("_")[0]

        removed  = prev_ncts - cur_ncts
        added    = cur_ncts  - prev_ncts

        # Fetch current status for removed NCTs
        removed_status: dict[str, tuple[str, str]] = {}
        if removed:
            for nct in removed:
                try:
                    r = requests.get(
                        f"https://clinicaltrials.gov/api/v2/studies/{nct}",
                        params={"fields": "protocolSection.statusModule,"
                                          "protocolSection.identificationModule"},
                        timeout=10,
                    )
                    d = r.json().get("protocolSection", {})
                    status  = d.get("statusModule", {}).get("overallStatus", "UNKNOWN")
                    why     = d.get("statusModule", {}).get("whyStopped", "")
                    title   = (d.get("identificationModule", {}).get("briefTitle", "")
                               or d.get("identificationModule", {}).get("officialTitle", ""))
                    removed_status[nct] = (status, why, title)
                except Exception:
                    removed_status[nct] = ("LOOKUP_FAILED", "", "")

        ws_ch.append(["Status", "NCT ID", "Trial Title", "Cancer Type",
                       "Phase", "Institution(s)", "Note"])
        for cell in ws_ch[1]:
            cell.font = hdr_font; cell.fill = navy_fill; cell.alignment = center_al

        # Removed rows
        for nct in sorted(removed):
            prev_row = df_prev[df_prev["NCT ID"] == nct].iloc[0] if not df_prev[df_prev["NCT ID"] == nct].empty else {}
            api_status, why, api_title = removed_status.get(nct, ("UNKNOWN", "", ""))
            title = prev_row.get("Study title", "") if hasattr(prev_row, "get") else api_title
            ct    = prev_row.get("Cancer type", "") if hasattr(prev_row, "get") else ""
            phase = prev_row.get("Phase", "") if hasattr(prev_row, "get") else ""
            if api_status in ("COMPLETED", "TERMINATED", "WITHDRAWN", "SUSPENDED", "ACTIVE_NOT_RECRUITING"):
                note = f"Status changed to {api_status}" + (f": {why}" if why else "")
            elif api_status == "RECRUITING":
                note = "Still RECRUITING — may have dropped SoCal site or geography shift. Investigate."
            else:
                note = f"Current API status: {api_status}"
            ws_ch.append(["REMOVED", nct, title or api_title, ct, phase, "", note])

        # Added rows
        for nct in sorted(added):
            rec = next((r for r in study_records if r["NCT ID"] == nct), {})
            inst = ", ".join(sorted({sr["Institution"] for sr in site_rows if sr["NCT ID"] == nct}))
            ws_ch.append(["NEW", nct, rec.get("Study title", ""),
                           rec.get("Cancer type", ""), rec.get("Phase", ""), inst,
                           f"New since {prev_date}"])

        ws_ch.column_dimensions["A"].width = 10
        ws_ch.column_dimensions["B"].width = 14
        ws_ch.column_dimensions["C"].width = 58
        ws_ch.column_dimensions["D"].width = 20
        ws_ch.column_dimensions["E"].width = 10
        ws_ch.column_dimensions["F"].width = 40
        ws_ch.column_dimensions["G"].width = 55

        # Colour-code rows
        red_fill   = PatternFill("solid", fgColor="FFCCCC")
        green_fill = PatternFill("solid", fgColor="C6EFCE")
        for row in ws_ch.iter_rows(min_row=2):
            status_val = row[0].value or ""
            fill = red_fill if status_val == "REMOVED" else (green_fill if status_val == "NEW" else None)
            if fill:
                for cell in row:
                    cell.fill = fill
            for cell in row:
                cell.font = body_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws_ch.row_dimensions[1].height = 22

    # ── Sheet 2: Stats by Hospital ─────────────────────────────────────────────
    ws_st.sheet_properties.tabColor = _C["teal"]
    cancer_groups = [
        ("Prostate",           ["Prostate"]),
        ("Bladder/Urothelial", ["Bladder/Urothelial"]),
        ("Kidney/Adrenal",     ["Kidney/RCC", "Adrenal"]),
        ("Testicular/GCT",     ["Testicular/GCT"]),
    ]
    inst_list: list[str] = []
    if not df_sites.empty and "Institution" in df_sites.columns:
        inst_list = sorted(df_sites["Institution"].unique().tolist())

    hdr_row = ["Institution", "Total"] + [cg[0] for cg in cancer_groups] + ["Phase 1", "Phase 2", "Phase 3", "Recruiting"]
    ws_st.append(hdr_row)
    for cell in ws_st[1]:
        cell.font = hdr_font; cell.fill = navy_fill; cell.alignment = center_al
    ws_st.freeze_panes = "A2"

    for inst in inst_list:
        inst_sites = df_sites[df_sites["Institution"] == inst] if not df_sites.empty else pd.DataFrame()
        inst_ncts  = set(inst_sites["NCT ID"].tolist()) if not inst_sites.empty else set()

        def ct_count(filters):
            if df_cur.empty or "NCT ID" not in df_cur.columns:
                return 0
            return sum(
                1 for nct in inst_ncts
                if any(ct in (df_cur[df_cur["NCT ID"] == nct]["Cancer type"].values[0]
                              if nct in df_cur["NCT ID"].values else "")
                       for ct in filters)
            )

        def phase_count(ph_str):
            if df_cur.empty or "NCT ID" not in df_cur.columns:
                return 0
            return sum(
                1 for nct in inst_ncts
                if ph_str in (df_cur[df_cur["NCT ID"] == nct]["Phase"].values[0]
                              if nct in df_cur["NCT ID"].values else "")
            )

        total  = len(inst_ncts)
        counts = [ct_count(cg[1]) for cg in cancer_groups]
        ph1    = phase_count("PHASE1")
        ph2    = phase_count("PHASE2")
        ph3    = phase_count("PHASE3")
        rec    = len(inst_sites[inst_sites["Status"] == "RECRUITING"]) if not inst_sites.empty and "Status" in inst_sites.columns else total

        row = [inst, total] + counts + [ph1, ph2, ph3, rec]
        ws_st.append(row)

    for ci, col in enumerate(ws_st.columns, 1):
        mx = max((len(str(c.value or "")) for c in col), default=8)
        ws_st.column_dimensions[get_column_letter(ci)].width = min(mx + 3, 40)

    # Style body
    alt_fill = PatternFill("solid", fgColor=_C["alt"])
    for ri, row in enumerate(ws_st.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.font = body_font
            if ri % 2 == 0:
                cell.fill = alt_fill
            cell.alignment = Alignment(horizontal="center" if cell.column > 1 else "left",
                                        vertical="top")

    # ── Sheet 3: New Trials by Window ─────────────────────────────────────────
    ws_nw.sheet_properties.tabColor = _C["green"]
    ws_nw.append(["Institution", "Cancer Type",
                  f"New since last pull ({run_date})" if prev_csv else "Total (baseline)",
                  f"New since 6 mo ({six_mo})",
                  f"New since 12 mo ({twelve_mo})"])
    for cell in ws_nw[1]:
        cell.font = hdr_font; cell.fill = navy_fill; cell.alignment = center_al
    ws_nw.freeze_panes = "A2"

    prev_ncts_set: set[str] = set()
    if prev_csv and prev_csv.exists():
        df_prev_nw = pd.read_csv(prev_csv, dtype=str).fillna("")
        prev_ncts_set = set(df_prev_nw["NCT ID"].tolist()) if "NCT ID" in df_prev_nw.columns else set()

    fpd_map: dict[str, str] = {}
    if "NCT ID" in df_cur.columns and "Study first posted" in df_cur.columns:
        for _, row in df_cur.iterrows():
            fpd_map[row["NCT ID"]] = str(row.get("Study first posted", "") or "")

    for inst in inst_list:
        inst_sites = df_sites[df_sites["Institution"] == inst] if not df_sites.empty else pd.DataFrame()
        inst_ncts  = set(inst_sites["NCT ID"].tolist()) if not inst_sites.empty else set()

        for cg_name, cg_filters in cancer_groups + [("All GU", ["Prostate","Bladder/Urothelial","Kidney/RCC","Adrenal","Testicular/GCT"])]:
            ct_ncts = {
                nct for nct in inst_ncts
                if any(ct in (df_cur[df_cur["NCT ID"] == nct]["Cancer type"].values[0]
                              if nct in df_cur["NCT ID"].values else "")
                       for ct in cg_filters)
            }
            if not ct_ncts:
                continue

            new_since_pull  = sum(1 for n in ct_ncts if n not in prev_ncts_set)
            new_since_6mo   = sum(1 for n in ct_ncts if fpd_map.get(n, "") >= six_mo)
            new_since_12mo  = sum(1 for n in ct_ncts if fpd_map.get(n, "") >= twelve_mo)

            ws_nw.append([inst, cg_name,
                          new_since_pull if prev_csv else len(ct_ncts),
                          new_since_6mo, new_since_12mo])

    for ci, col in enumerate(ws_nw.columns, 1):
        mx = max((len(str(c.value or "")) for c in col), default=8)
        ws_nw.column_dimensions[get_column_letter(ci)].width = min(mx + 3, 50)

    for ri, row in enumerate(ws_nw.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.font = body_font
            if ri % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=_C["alt"])
            cell.alignment = Alignment(horizontal="center" if cell.column > 2 else "left",
                                        vertical="top")

    # ── Sheet 4: Metadata ─────────────────────────────────────────────────────
    ws_me.sheet_properties.tabColor = _C["gold"]
    recall_info = ""
    if recall_result:
        found = recall_result.get("found", 0)
        total = recall_result.get("total", 0)
        missed_recruiting = recall_result.get("missed_recruiting", [])
        recall_info = f"{found}/{total} ({100*found//total if total else 0}%)"
        if missed_recruiting:
            recall_info += f" — {len(missed_recruiting)} still-RECRUITING missed (see Step 3 output)"

    meta_rows = [
        ("Field",                     "Value"),
        ("Script",                    f"gu_pipeline.py v{VERSION}"),
        ("Run date",                  run_ts.split("_")[0]),
        ("Run time",                  run_ts.split("_")[1]),
        ("Pulled by",                 PULLED_BY),
        ("",                          ""),
        ("── Data source ──",          ""),
        ("API",                       "ClinicalTrials.gov API v2"),
        ("API base URL",              BASE_URL),
        ("",                          ""),
        ("── NCCN taxonomy versions ──",""),
        ("Prostate",                  get_nccn_stamp("Prostate")),
        ("Bladder/Urothelial",        get_nccn_stamp("Bladder/Urothelial")),
        ("Kidney/RCC",                get_nccn_stamp("Kidney/RCC")),
        ("Testicular/GCT",            get_nccn_stamp("Testicular/GCT")),
        ("Adrenal",                   "Not yet classified — taxonomy build pending"),
        ("",                          ""),
        ("── Filters ──",              ""),
        ("Study type",                STUDY_TYPE),
        ("Status",                    ", ".join(STATUSES)),
        ("Geography",                 GEO_FILTER),
        ("Conditions queried",        str(len(CONDITIONS))),
        *[(f"  Condition {i+1}",       c) for i, c in enumerate(CONDITIONS)],
        ("Target institutions",       str(len(TARGET_INSTITUTIONS))),
        ("",                          ""),
        ("── Results ──",              ""),
        ("Unique trials",             str(len(study_records))),
        ("Site × institution rows",   str(len(site_rows))),
        ("Institutions found",        str(len({r["Institution"] for r in site_rows}) if site_rows else 0)),
        ("",                          ""),
        ("── Validation ──",           ""),
        ("UCI/Hoag recall (Step 3)",  recall_info or "Not recorded this run"),
        ("Previous pull CSV",         str(prev_csv) if prev_csv else "None (first run / baseline)"),
    ]

    for i, (k, v) in enumerate(meta_rows, start=1):
        ws_me.cell(i, 1, k)
        ws_me.cell(i, 2, v)
        if i == 1:
            ws_me.cell(i, 1).font = hdr_font; ws_me.cell(i, 1).fill = gold_fill
            ws_me.cell(i, 2).font = hdr_font; ws_me.cell(i, 2).fill = gold_fill
        elif k.startswith("──") or k == "":
            ws_me.cell(i, 1).font = Font(bold=True, color="888888", name="Arial", size=9, italic=True)
        else:
            ws_me.cell(i, 1).font = key_font
            ws_me.cell(i, 2).font = body_font

    ws_me.column_dimensions["A"].width = 30
    ws_me.column_dimensions["B"].width = 60

    wb.save(out_path)
    print(f"  {out_path.name}")


# ── Main ──────────────────────────────────────────────────────────────────────
def main() -> None:
    run_ts   = datetime.now().strftime("%Y-%m-%d_%H%M")
    run_date = run_ts.split("_")[0]

    # Create dated output subfolder
    run_out = OUT / run_ts
    run_out.mkdir(parents=True, exist_ok=True)

    w = 62
    print("=" * w)
    print(f"  GU ONCOLOGY SOCAL — v{VERSION}")
    print(f"  Run: {run_ts}   Pulled by: {PULLED_BY}")
    print(f"  Filters : RECRUITING · INTERVENTIONAL · SoCal 180mi")
    print(f"  Output  : {run_out}")
    print("=" * w + "\n")

    # ── Find previous CSV for delta ────────────────────────────────────────────
    prev_csv = _find_prev_csv(OUT, run_ts)
    if prev_csv:
        print(f"  Previous pull found: {prev_csv.parent.name}/")
    else:
        print("  No previous pull found — this run will be the baseline.")

    # ── Fetch ──────────────────────────────────────────────────────────────────
    t0      = time.time()
    studies = fetch_all()
    print(f"\n  Fetch complete in {time.time()-t0:.0f}s")

    # ── Parse ──────────────────────────────────────────────────────────────────
    print("\nParsing studies…")
    study_records: list[dict] = []
    site_rows:     list[dict] = []
    seen_ncts:      set[str]  = set()
    seen_site_keys: set[tuple]= set()

    for study in studies:
        record, s_rows = parse_study_extended(study)
        if record is None:
            continue
        nct = record["NCT ID"]
        if nct not in seen_ncts:
            seen_ncts.add(nct)
            study_records.append(record)
        for row in s_rows:
            key = (row["NCT ID"], row["Institution"])
            if key not in seen_site_keys:
                seen_site_keys.add(key)
                site_rows.append(row)

    if not study_records:
        print("No qualifying trials found.")
        sys.exit(1)

    print(f"\n{'=' * w}")
    print(f"  Unique trials            : {len(study_records)}")
    print(f"  Site × institution rows  : {len(site_rows)}")
    unique_insts = len({r['Institution'] for r in site_rows})
    print(f"  Institutions             : {unique_insts}")
    basket_n = sum(1 for r in study_records if r["Cancer type"] == "Basket")
    print(f"  GU-specific trials       : {len(study_records) - basket_n}")
    print(f"  Basket (residual)        : {basket_n}")
    print(f"{'=' * w}\n")

    # ── Disease classification (all 4 NCCN cancer types) ──────────────────────
    print("Classifying disease settings (NCCN)…")
    nct_classification: dict[str, dict] = {}
    conf_tally: dict[str, dict[str, int]] = {
        "Prostate": {"HIGH": 0, "MEDIUM": 0, "LOW": 0, "N/A": 0, "UNCLASSIFIED": 0},
        "Bladder/Urothelial": {"HIGH": 0, "MEDIUM": 0, "LOW": 0, "N/A": 0, "UNCLASSIFIED": 0},
        "Kidney/RCC": {"HIGH": 0, "MEDIUM": 0, "LOW": 0, "N/A": 0, "UNCLASSIFIED": 0},
        "Testicular/GCT": {"HIGH": 0, "MEDIUM": 0, "LOW": 0, "N/A": 0, "UNCLASSIFIED": 0},
    }

    for rec in study_records:
        cancer_types = rec.get("Cancer type", "")
        primary_ct   = cancer_types.split(" | ")[0] if " | " in cancer_types else cancer_types

        clf = classify_trial(
            cancer_type      = primary_ct,
            title            = rec.get("Study title", ""),
            eligibility_incl = rec.get("_inclusion_raw", ""),
            eligibility_excl = rec.get("_exclusion_raw", ""),
            conditions       = rec.get("Conditions", ""),
            interventions    = rec.get("_interventions_raw", ""),
            brief_summary    = rec.get("Brief summary", ""),
        )

        rec["Disease setting (primary)"]  = clf.disease_setting_primary
        rec["Disease setting (all)"]      = clf.disease_setting_all
        rec["Classification confidence"]  = clf.classification_confidence
        rec["BCG status"]                 = clf.bcg_status
        rec["Cisplatin status"]           = clf.cisplatin_status
        rec["Treatment modality"]         = clf.treatment_modality_str
        rec["Is combination"]             = clf.is_combination
        rec["Delivery"]                   = clf.delivery
        rec["NCCN taxonomy version"]      = (
            f"NCCN v{clf.nccn_version} ({clf.nccn_date})" if clf.nccn_version else "N/A"
        )

        if primary_ct in conf_tally:
            ck = clf.classification_confidence
            conf_tally[primary_ct][ck if ck in conf_tally[primary_ct] else "UNCLASSIFIED"] += 1

        nct_classification[rec["NCT ID"]] = {
            "Disease setting (primary)": clf.disease_setting_primary,
            "Classification confidence": clf.classification_confidence,
            "Treatment modality":        clf.treatment_modality_str,
            "Delivery":                  clf.delivery,
        }

    # Backfill classification into site_rows
    for sr in site_rows:
        clf_data = nct_classification.get(sr["NCT ID"], {})
        sr["Disease setting (primary)"] = clf_data.get("Disease setting (primary)", "")
        sr["Classification confidence"] = clf_data.get("Classification confidence", "")
        sr["Treatment modality"]        = clf_data.get("Treatment modality", "")
        sr["Delivery"]                  = clf_data.get("Delivery", "")

    # Remove internal raw fields
    for rec in study_records:
        rec.pop("_interventions_raw", None)
        rec.pop("_inclusion_raw", None)
        rec.pop("_exclusion_raw", None)

    # Print classification summary
    for ct, tally in conf_tally.items():
        total_ct = sum(tally.values())
        if total_ct > 0:
            print(f"\n  {ct} ({total_ct} trials):")
            for level, n in tally.items():
                if n > 0:
                    print(f"    {level:<14}: {n:>3}")

    # ── Save flat CSV (used for delta on next run) ────────────────────────────
    csv_cols = [
        "NCT ID", "Study title", "Phase", "Status", "Cancer type",
        "Disease setting (primary)", "Disease setting (all)",
        "Classification confidence", "BCG status", "Cisplatin status",
        "Treatment modality", "Is combination", "Delivery",
        "NCCN taxonomy version",
        "Conditions", "Lead sponsor", "Additional sponsors",
        "Information provided by", "Intervention(s)", "Study design",
        "Enrollment (estimated)", "Enrollment (actual)", "Other study IDs",
        "Start date", "Primary completion", "Last update posted", "Study first posted",
        "Last pulled by", "ClinicalTrials URL",
    ]
    df_csv_all = pd.DataFrame(study_records)
    df_csv = df_csv_all[[c for c in csv_cols if c in df_csv_all.columns]]
    csv_out = run_out / f"all_trials_{run_date}.csv"
    df_csv.to_csv(csv_out, index=False)
    print(f"\n  Flat CSV saved → {csv_out.name}  ({len(df_csv)} rows)")

    # ── Write 4 cancer-type Excel files ──────────────────────────────────────
    print(f"\nBuilding Excel files…")
    cancer_type_defs = [
        ("prostate",       "Prostate",        ["Prostate"]),
        ("bladder",        "Bladder/Urothelial", ["Bladder/Urothelial"]),
        ("kidney_adrenal", "Kidney/RCC",      ["Kidney/RCC", "Adrenal"]),
        ("testicular",     "Testicular/GCT",  ["Testicular/GCT"]),
    ]
    for fname, nccn_ct, ct_filters in cancer_type_defs:
        out_path = run_out / f"{fname}_trials_{run_date}.xlsx"
        print(f"\n  → {out_path.name}")
        build_cancer_type_excel(
            cancer_type=nccn_ct,
            cancer_type_filters=ct_filters,
            study_records=study_records,
            site_rows=site_rows,
            run_ts=run_ts,
            out_path=out_path,
        )

    # ── Write updates file ────────────────────────────────────────────────────
    updates_out = run_out / f"updates_{run_date}.xlsx"
    print(f"\n  → {updates_out.name}")
    build_updates_excel(
        study_records=study_records,
        site_rows=site_rows,
        run_ts=run_ts,
        prev_csv=prev_csv,
        out_path=updates_out,
    )

    # ── Top PIs summary ───────────────────────────────────────────────────────
    df_s = pd.DataFrame(site_rows)
    df_s["canon"] = df_s["PI name"].apply(canonical_pi)
    top = (
        df_s[df_s["canon"] != ""]
        .groupby("canon")
        .agg(name=("PI name", lambda x: x.value_counts().index[0]),
             n   =("NCT ID",  "nunique"))
        .sort_values("n", ascending=False)
        .head(10)
    )
    print(f"\n── TOP 10 PIs " + "─" * (w - 13))
    for _, r in top.iterrows():
        print(f"  {r['name']:<38} {int(r['n']):>3} trials")

    print(f"\n{'=' * w}")
    print(f"  Output folder: {run_out}")
    print(f"{'=' * w}")


if __name__ == "__main__":
    main()
