from __future__ import annotations

from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from .nccn_classifier import get_disease_settings_in_order, get_nccn_stamp, get_section_colors
    from .site_normalization import canonical_pi
except ImportError:
    from nccn_classifier import get_disease_settings_in_order, get_nccn_stamp, get_section_colors
    from site_normalization import canonical_pi


_COLORS = {
    "navy": "1F4E79",
    "blue": "2E75B6",
    "red": "C00000",
    "green": "375623",
    "teal": "1F6B75",
    "gold": "7B5C00",
    "white": "FFFFFF",
    "link": "0563C1",
    "alt": "DCE6F1",
    "dark_gray": "404040",
    "light_gray": "F2F2F2",
    "conf_high": "C6EFCE",
    "conf_medium": "FFEB9C",
    "conf_low": "FFCC99",
    "map_zero": "F2F2F2",
    "map_low": "D6E4F0",
    "map_mid": "2E6DA4",
    "map_high": "1B3A6B",
}

_HEADER_FONT = Font(bold=True, color=_COLORS["white"], name="Arial", size=10)
_BODY_FONT = Font(name="Arial", size=10)
_LINK_FONT = Font(color=_COLORS["link"], underline="single", name="Arial", size=10)
_HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
_WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")


def find_prev_csv(out_root: Path, current_ts: str) -> Path | None:
    candidates: list[Path] = []
    for directory in out_root.iterdir():
        if not directory.is_dir():
            continue
        if directory.name >= current_ts:
            continue
        candidates.extend(directory.glob("all_trials_*.csv"))
    return max(candidates, key=lambda path: path.parent.name) if candidates else None


def save_flat_csv(study_records: list[dict], out_path: Path) -> int:
    columns = [
        "NCT ID", "Study title", "Phase", "Status", "Cancer type",
        "Disease setting (primary)", "Disease setting (all)",
        "Classification confidence", "BCG status", "Cisplatin status",
        "Treatment modality", "Is combination", "Delivery",
        "NCCN taxonomy version", "Conditions", "Lead sponsor",
        "Additional sponsors", "Information provided by", "Intervention(s)",
        "Study design", "Enrollment (estimated)", "Enrollment (actual)",
        "Other study IDs", "Start date", "Primary completion",
        "Last update posted", "Study first posted", "Last pulled by",
        "ClinicalTrials URL",
    ]
    frame = pd.DataFrame(study_records)
    trimmed = frame[[column for column in columns if column in frame.columns]]
    trimmed.to_csv(out_path, index=False)
    return len(trimmed)


def _style_ws(
    worksheet,
    tab_color: str,
    *,
    url_col: str | None = None,
    wrap_cols: list[str] | None = None,
    freeze: str = "A2",
    row_h: int = 15,
) -> None:
    wrap_set = set(wrap_cols or [])
    url_col_no = None

    for cell in worksheet[1]:
        cell.font = _HEADER_FONT
        cell.fill = PatternFill("solid", fgColor=tab_color)
        cell.alignment = _HEADER_ALIGN
        if cell.value == url_col:
            url_col_no = cell.column

    worksheet.row_dimensions[1].height = 28
    worksheet.freeze_panes = freeze
    worksheet.sheet_properties.tabColor = tab_color

    for row_index, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        is_alt_row = row_index % 2 == 0
        for cell in row:
            is_link = url_col_no and cell.column == url_col_no
            header = worksheet.cell(1, cell.column).value or ""
            cell.font = _LINK_FONT if is_link else _BODY_FONT
            if is_alt_row:
                cell.fill = PatternFill("solid", fgColor=_COLORS["alt"])
            cell.alignment = _WRAP_ALIGN if header in wrap_set else Alignment(vertical="top")
            if is_link and cell.value:
                cell.hyperlink = str(cell.value)
        worksheet.row_dimensions[row_index].height = row_h

    for col_index, column in enumerate(worksheet.columns, start=1):
        header = worksheet.cell(1, col_index).value or ""
        max_len = max((len(str(cell.value or "")) for cell in column), default=10)
        worksheet.column_dimensions[get_column_letter(col_index)].width = (
            70 if header in wrap_set else min(max_len + 3, 55)
        )


def _df_to_ws(worksheet, frame: pd.DataFrame) -> None:
    worksheet.append(list(frame.columns))
    for row in frame.itertuples(index=False):
        worksheet.append([None if isinstance(value, float) and pd.isna(value) else value for value in row])


def _phase_sort_key(phase: str) -> int:
    phase_upper = (phase or "").upper()
    if "3" in phase_upper:
        return 4
    if "2" in phase_upper and "3" in phase_upper:
        return 3
    if "2" in phase_upper:
        return 3
    if "1" in phase_upper and "2" in phase_upper:
        return 2
    if "1" in phase_upper:
        return 1
    return 0


def _build_trial_finder(
    worksheet,
    study_records: list[dict],
    site_rows: list[dict],
    *,
    cancer_type: str,
    cancer_type_filters: list[str] | None = None,
) -> tuple[int, int]:
    if cancer_type_filters is None:
        cancer_type_filters = [cancer_type]

    taxonomy_colors = get_section_colors(cancer_type)
    palette = ["1F4E79", "2E75B6", "375623", "1F6B75", "7B5C00", "C00000", "404040", "1B3A6B"]
    palette_index = [0]

    def section_color(section: str) -> str:
        if section in taxonomy_colors:
            return taxonomy_colors[section]
        color = palette[palette_index[0] % len(palette)]
        palette_index[0] += 1
        return color

    headers = ["NCT ID", "Trial Title", "Phase", "Treatment", "Available At", "Top PI(s)", "Confidence", "Link"]
    widths = [14, 55, 8, 28, 45, 32, 11, 8]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width

    nct_to_institutions: dict[str, set[str]] = defaultdict(set)
    nct_to_pis: dict[str, list[str]] = defaultdict(list)
    nct_to_pi_email: dict[str, dict[str, str]] = defaultdict(dict)
    seen_pi_keys: dict[str, set[str]] = defaultdict(set)

    for site_row in site_rows:
        nct_id = site_row["NCT ID"]
        nct_to_institutions[nct_id].add(site_row["Institution"])
        pi_name = site_row.get("PI name", "")
        if pi_name and pi_name != "Not listed":
            canon = canonical_pi(pi_name)
            if canon and canon not in seen_pi_keys[nct_id]:
                seen_pi_keys[nct_id].add(canon)
                nct_to_pis[nct_id].append(pi_name)
            email = site_row.get("PI email", "")
            if canon and email and canon not in nct_to_pi_email[nct_id]:
                nct_to_pi_email[nct_id][canon] = email

    ordered_institutions = {nct_id: sorted(values) for nct_id, values in nct_to_institutions.items()}
    disease_settings = get_disease_settings_in_order(cancer_type)

    link_font = Font(color=_COLORS["link"], underline="single", name="Arial", size=9)
    body_font = Font(name="Arial", size=9)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    left_mid = Alignment(horizontal="left", vertical="center")
    confidence_fills = {
        "HIGH": PatternFill("solid", fgColor=_COLORS["conf_high"]),
        "MEDIUM": PatternFill("solid", fgColor=_COLORS["conf_medium"]),
        "LOW": PatternFill("solid", fgColor=_COLORS["conf_low"]),
    }
    confidence_fonts = {
        "HIGH": Font(name="Arial", size=9, color="276221"),
        "MEDIUM": Font(name="Arial", size=9, color="7A5C00"),
        "LOW": Font(name="Arial", size=9, color="8B3000"),
    }

    current_row = 1
    last_section = None
    settings_written = 0
    entries_written = 0

    for disease_setting in disease_settings:
        label = disease_setting["label"]
        section = disease_setting["section"]
        matching = [
            record for record in study_records
            if any(cancer in record.get("Cancer type", "") for cancer in cancer_type_filters)
            and label in record.get("Disease setting (all)", "")
        ]
        if not matching:
            continue

        settings_written += 1
        sec_color = section_color(section)
        if section != last_section:
            last_section = section
            worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
            cell = worksheet.cell(current_row, 1, section.upper())
            cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
            cell.fill = PatternFill("solid", fgColor=sec_color)
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            worksheet.row_dimensions[current_row].height = 22
            current_row += 1

        worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
        lighter = "".join(f"{min(int(sec_color[index:index + 2], 16) + 30, 255):02X}" for index in (0, 2, 4))
        label_text = f"{label}  —  {len(matching)} trial{'s' if len(matching) != 1 else ''}"
        cell = worksheet.cell(current_row, 1, label_text)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = PatternFill("solid", fgColor=lighter)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        worksheet.row_dimensions[current_row].height = 20
        current_row += 1

        for index, header in enumerate(headers, start=1):
            cell = worksheet.cell(current_row, index, header)
            cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
            cell.fill = PatternFill("solid", fgColor=_COLORS["dark_gray"])
            cell.alignment = center
        worksheet.row_dimensions[current_row].height = 18
        current_row += 1

        matching_sorted = sorted(
            matching,
            key=lambda record: (
                -_phase_sort_key(record.get("Phase", "")),
                -len(ordered_institutions.get(record["NCT ID"], [])),
            ),
        )

        for row_index, record in enumerate(matching_sorted):
            nct_id = record["NCT ID"]
            ctgov_url = f"https://clinicaltrials.gov/study/{nct_id}"
            alt_fill = PatternFill("solid", fgColor="F7F7F7") if row_index % 2 == 1 else None
            pi_parts = []
            for pi_name in nct_to_pis.get(nct_id, [])[:3]:
                canon = canonical_pi(pi_name)
                email = nct_to_pi_email.get(nct_id, {}).get(canon, "")
                pi_parts.append(f"{pi_name} ({email})" if email else pi_name)

            values = [
                nct_id,
                (record.get("Study title") or "")[:75],
                record.get("Phase", "N/A"),
                record.get("Treatment modality", ""),
                ", ".join(ordered_institutions.get(nct_id, [])),
                "; ".join(pi_parts),
                record.get("Classification confidence", ""),
                "→ CT.gov",
            ]

            for column, value in enumerate(values, start=1):
                cell = worksheet.cell(current_row, column, value)
                if column in (1, 8):
                    cell.font = link_font
                    cell.hyperlink = ctgov_url
                    cell.alignment = center if column == 8 else left_mid
                elif column == 3:
                    cell.font = body_font
                    cell.alignment = center
                elif column == 7:
                    confidence = record.get("Classification confidence", "")
                    cell.font = confidence_fonts.get(confidence, body_font)
                    cell.fill = confidence_fills.get(confidence, PatternFill())
                    cell.alignment = center
                else:
                    cell.font = body_font
                    cell.alignment = left
                if alt_fill and column != 7:
                    cell.fill = alt_fill

            worksheet.row_dimensions[current_row].height = 16
            current_row += 1
            entries_written += 1

        worksheet.row_dimensions[current_row].height = 6
        current_row += 1

    return settings_written, entries_written


def build_cancer_type_excel(
    *,
    cancer_type: str,
    cancer_type_filters: list[str],
    study_records: list[dict],
    site_rows: list[dict],
    out_path: Path,
) -> None:
    cancer_records = [
        record for record in study_records
        if any(cancer in record.get("Cancer type", "") for cancer in cancer_type_filters)
    ]
    cancer_sites = [
        row for row in site_rows
        if any(cancer in row.get("Cancer type", "") for cancer in cancer_type_filters)
    ]
    if not cancer_records:
        return

    records_frame = pd.DataFrame(cancer_records)
    sites_frame = pd.DataFrame(cancer_sites)
    sites_columns = [
        "NCT ID", "Trial title", "Phase", "Status", "Cancer type",
        "Disease setting (primary)", "Classification confidence",
        "Treatment modality", "Delivery", "Institution", "PI name",
        "PI email", "PI phone", "PI affiliation", "Site city", "Lead sponsor",
    ]
    eligibility_columns = [
        "NCT ID", "Study title", "Cancer type", "Disease setting (primary)",
        "Phase", "Min age", "Max age", "Sex", "Inclusion criteria", "Exclusion criteria",
    ]
    outcomes_columns = [
        "NCT ID", "Study title", "Cancer type", "Disease setting (primary)",
        "Phase", "Primary outcomes", "Secondary outcomes",
    ]

    sites_output = (
        sites_frame[[column for column in sites_columns if column in sites_frame.columns]]
        .copy()
        .drop_duplicates(["NCT ID", "Institution"])
    )
    eligibility_output = records_frame[[column for column in eligibility_columns if column in records_frame.columns]].copy()
    outcomes_output = records_frame[[column for column in outcomes_columns if column in records_frame.columns]].copy()

    workbook = Workbook()
    trial_finder = workbook.active
    trial_finder.title = "Trial Finder"
    sites_sheet = workbook.create_sheet("Sites × PI")
    eligibility_sheet = workbook.create_sheet("Eligibility")
    outcomes_sheet = workbook.create_sheet("Outcomes")

    _build_trial_finder(
        trial_finder,
        cancer_records,
        cancer_sites,
        cancer_type=cancer_type,
        cancer_type_filters=cancer_type_filters,
    )
    _df_to_ws(sites_sheet, sites_output)
    _df_to_ws(eligibility_sheet, eligibility_output)
    _df_to_ws(outcomes_sheet, outcomes_output)

    _style_ws(sites_sheet, _COLORS["blue"], row_h=16)
    _style_ws(eligibility_sheet, _COLORS["red"], wrap_cols=["Inclusion criteria", "Exclusion criteria"], row_h=60)
    _style_ws(outcomes_sheet, _COLORS["green"], wrap_cols=["Primary outcomes", "Secondary outcomes"], row_h=45)
    workbook.save(out_path)


def build_updates_excel(
    *,
    study_records: list[dict],
    site_rows: list[dict],
    run_ts: str,
    prev_csv: Path | None,
    out_path: Path,
    client=None,
    api_base_url: str,
    conditions: list[str],
    target_institutions: list[str],
    version: str,
    pulled_by: str,
) -> None:
    from datetime import datetime, timedelta

    run_date = run_ts.split("_")[0]
    today = datetime.strptime(run_date, "%Y-%m-%d")
    six_months = (today - timedelta(days=183)).strftime("%Y-%m-%d")
    twelve_months = (today - timedelta(days=365)).strftime("%Y-%m-%d")

    current_frame = pd.DataFrame(study_records)
    sites_frame = pd.DataFrame(site_rows) if site_rows else pd.DataFrame()

    workbook = Workbook()
    changes_sheet = workbook.active
    changes_sheet.title = "Changes since last pull"
    stats_sheet = workbook.create_sheet("Stats by Hospital")
    windows_sheet = workbook.create_sheet("New Trials by Window")
    metadata_sheet = workbook.create_sheet("Metadata")

    gold_fill = PatternFill("solid", fgColor=_COLORS["gold"])
    navy_fill = PatternFill("solid", fgColor=_COLORS["navy"])
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    body_font = Font(name="Arial", size=10)
    key_font = Font(bold=True, name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")

    current_ncts = set(current_frame["NCT ID"].tolist()) if "NCT ID" in current_frame.columns else set()
    if prev_csv is None or not prev_csv.exists():
        changes_sheet.append(["Status", "NCT ID", "Trial Title", "Cancer Type", "Phase", "Institution(s)", "Note"])
        for cell in changes_sheet[1]:
            cell.font = header_font
            cell.fill = navy_fill
            cell.alignment = center
        changes_sheet.append([
            "BASELINE", "", "First run — no prior pull available for comparison.",
            "", "", "", f"Baseline established {run_date}. All {len(current_ncts)} trials are new.",
        ])
    else:
        previous_frame = pd.read_csv(prev_csv, dtype=str).fillna("")
        previous_ncts = set(previous_frame["NCT ID"].tolist()) if "NCT ID" in previous_frame.columns else set()
        previous_date = prev_csv.parent.name.split("_")[0]
        removed = previous_ncts - current_ncts
        added = current_ncts - previous_ncts

        changes_sheet.append(["Status", "NCT ID", "Trial Title", "Cancer Type", "Phase", "Institution(s)", "Note"])
        for cell in changes_sheet[1]:
            cell.font = header_font
            cell.fill = navy_fill
            cell.alignment = center

        removed_status: dict[str, dict[str, str]] = {}
        if client is not None:
            for nct_id in sorted(removed):
                try:
                    removed_status[nct_id] = client.fetch_study_status(nct_id)
                except Exception:
                    removed_status[nct_id] = {"status": "LOOKUP_FAILED", "why_stopped": "", "title": ""}

        for nct_id in sorted(removed):
            previous_rows = previous_frame[previous_frame["NCT ID"] == nct_id]
            previous_row = previous_rows.iloc[0] if not previous_rows.empty else {}
            status_info = removed_status.get(nct_id, {"status": "UNKNOWN", "why_stopped": "", "title": ""})
            title = previous_row.get("Study title", "") if hasattr(previous_row, "get") else status_info["title"]
            cancer_type = previous_row.get("Cancer type", "") if hasattr(previous_row, "get") else ""
            phase = previous_row.get("Phase", "") if hasattr(previous_row, "get") else ""
            status = status_info["status"]
            why_stopped = status_info["why_stopped"]
            if status in {"COMPLETED", "TERMINATED", "WITHDRAWN", "SUSPENDED", "ACTIVE_NOT_RECRUITING"}:
                note = f"Status changed to {status}" + (f": {why_stopped}" if why_stopped else "")
            elif status == "RECRUITING":
                note = "Still RECRUITING — investigate geography, site normalization, or query coverage."
            else:
                note = f"Current API status: {status}"
            changes_sheet.append(["REMOVED", nct_id, title or status_info["title"], cancer_type, phase, "", note])

        for nct_id in sorted(added):
            record = next((item for item in study_records if item["NCT ID"] == nct_id), {})
            institutions = ", ".join(sorted({row["Institution"] for row in site_rows if row["NCT ID"] == nct_id}))
            changes_sheet.append([
                "NEW", nct_id, record.get("Study title", ""), record.get("Cancer type", ""),
                record.get("Phase", ""), institutions, f"New since {previous_date}",
            ])

    for index, column in enumerate(changes_sheet.columns, start=1):
        max_len = max((len(str(cell.value or "")) for cell in column), default=8)
        changes_sheet.column_dimensions[get_column_letter(index)].width = min(max_len + 3, 60)
    for row in changes_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    cancer_groups = [
        ("Prostate", ["Prostate"]),
        ("Bladder/Urothelial", ["Bladder/Urothelial"]),
        ("Kidney/Adrenal", ["Kidney/RCC", "Adrenal"]),
        ("Testicular/GCT", ["Testicular/GCT"]),
    ]
    institutions = sorted(sites_frame["Institution"].unique().tolist()) if not sites_frame.empty and "Institution" in sites_frame.columns else []
    stats_sheet.append(["Institution", "Total"] + [name for name, _ in cancer_groups] + ["Phase 1", "Phase 2", "Phase 3", "Recruiting"])
    for cell in stats_sheet[1]:
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = center
    stats_sheet.freeze_panes = "A2"

    for institution in institutions:
        institution_sites = sites_frame[sites_frame["Institution"] == institution]
        institution_ncts = set(institution_sites["NCT ID"].tolist())

        def cancer_count(filters: list[str]) -> int:
            if current_frame.empty or "NCT ID" not in current_frame.columns:
                return 0
            return sum(
                1 for nct_id in institution_ncts
                if any(
                    cancer in (
                        current_frame[current_frame["NCT ID"] == nct_id]["Cancer type"].values[0]
                        if nct_id in current_frame["NCT ID"].values else ""
                    )
                    for cancer in filters
                )
            )

        def phase_count(phase_token: str) -> int:
            if current_frame.empty or "NCT ID" not in current_frame.columns:
                return 0
            return sum(
                1 for nct_id in institution_ncts
                if phase_token in (
                    current_frame[current_frame["NCT ID"] == nct_id]["Phase"].values[0]
                    if nct_id in current_frame["NCT ID"].values else ""
                )
            )

        stats_sheet.append([
            institution,
            len(institution_ncts),
            *[cancer_count(filters) for _, filters in cancer_groups],
            phase_count("PHASE1"),
            phase_count("PHASE2"),
            phase_count("PHASE3"),
            len(institution_sites[institution_sites["Status"] == "RECRUITING"]) if "Status" in institution_sites.columns else len(institution_ncts),
        ])

    for index, column in enumerate(stats_sheet.columns, start=1):
        max_len = max((len(str(cell.value or "")) for cell in column), default=8)
        stats_sheet.column_dimensions[get_column_letter(index)].width = min(max_len + 3, 40)

    windows_sheet.append([
        "Institution",
        "Cancer Type",
        f"New since last pull ({run_date})" if prev_csv else "Total (baseline)",
        f"New since 6 mo ({six_months})",
        f"New since 12 mo ({twelve_months})",
    ])
    for cell in windows_sheet[1]:
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = center
    windows_sheet.freeze_panes = "A2"

    previous_ncts = set()
    if prev_csv and prev_csv.exists():
        previous_ncts = set(pd.read_csv(prev_csv, dtype=str).fillna("").get("NCT ID", []))
    first_posted_map = {}
    if "NCT ID" in current_frame.columns and "Study first posted" in current_frame.columns:
        for _, row in current_frame.iterrows():
            first_posted_map[row["NCT ID"]] = str(row.get("Study first posted", "") or "")

    for institution in institutions:
        institution_sites = sites_frame[sites_frame["Institution"] == institution]
        institution_ncts = set(institution_sites["NCT ID"].tolist())
        all_groups = cancer_groups + [("All GU", ["Prostate", "Bladder/Urothelial", "Kidney/RCC", "Adrenal", "Testicular/GCT"])]
        for group_name, filters in all_groups:
            group_ncts = {
                nct_id for nct_id in institution_ncts
                if any(
                    cancer in (
                        current_frame[current_frame["NCT ID"] == nct_id]["Cancer type"].values[0]
                        if nct_id in current_frame["NCT ID"].values else ""
                    )
                    for cancer in filters
                )
            }
            if not group_ncts:
                continue
            windows_sheet.append([
                institution,
                group_name,
                sum(1 for nct_id in group_ncts if nct_id not in previous_ncts) if prev_csv else len(group_ncts),
                sum(1 for nct_id in group_ncts if first_posted_map.get(nct_id, "") >= six_months),
                sum(1 for nct_id in group_ncts if first_posted_map.get(nct_id, "") >= twelve_months),
            ])

    for index, column in enumerate(windows_sheet.columns, start=1):
        max_len = max((len(str(cell.value or "")) for cell in column), default=8)
        windows_sheet.column_dimensions[get_column_letter(index)].width = min(max_len + 3, 50)

    metadata_rows = [
        ("Field", "Value"),
        ("Script", f"gu_pipeline.py v{version}"),
        ("Run date", run_date),
        ("Run time", run_ts.split("_")[1]),
        ("Pulled by", pulled_by),
        ("", ""),
        ("── Data source ──", ""),
        ("API", "ClinicalTrials.gov API v2"),
        ("API base URL", api_base_url),
        ("", ""),
        ("── NCCN taxonomy versions ──", ""),
        ("Prostate", get_nccn_stamp("Prostate")),
        ("Bladder/Urothelial", get_nccn_stamp("Bladder/Urothelial")),
        ("Kidney/RCC", get_nccn_stamp("Kidney/RCC")),
        ("Testicular/GCT", get_nccn_stamp("Testicular/GCT")),
        ("Adrenal", "Not yet classified — taxonomy build pending"),
        ("", ""),
        ("── Filters ──", ""),
        ("Conditions queried", str(len(conditions))),
        *[(f"  Condition {index + 1}", condition) for index, condition in enumerate(conditions)],
        ("Target institutions", str(len(target_institutions))),
        ("", ""),
        ("── Results ──", ""),
        ("Unique trials", str(len(study_records))),
        ("Site × institution rows", str(len(site_rows))),
        ("Institutions found", str(len({row['Institution'] for row in site_rows}) if site_rows else 0)),
        ("Previous pull CSV", str(prev_csv) if prev_csv else "None (baseline)"),
    ]

    for index, (key, value) in enumerate(metadata_rows, start=1):
        metadata_sheet.cell(index, 1, key)
        metadata_sheet.cell(index, 2, value)
        if index == 1:
            metadata_sheet.cell(index, 1).font = header_font
            metadata_sheet.cell(index, 1).fill = gold_fill
            metadata_sheet.cell(index, 2).font = header_font
            metadata_sheet.cell(index, 2).fill = gold_fill
        elif key.startswith("──") or key == "":
            metadata_sheet.cell(index, 1).font = Font(bold=True, color="888888", name="Arial", size=9, italic=True)
        else:
            metadata_sheet.cell(index, 1).font = key_font
            metadata_sheet.cell(index, 2).font = body_font

    metadata_sheet.column_dimensions["A"].width = 30
    metadata_sheet.column_dimensions["B"].width = 60
    workbook.save(out_path)
